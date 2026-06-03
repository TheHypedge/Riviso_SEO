import base64
import os
import re
import json
import uuid
import secrets
import threading
import time
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from io import BytesIO
import zipfile
from typing import Any
from urllib.parse import urlencode, urlparse
from urllib.request import urlopen

try:
    from zoneinfo import ZoneInfo, available_timezones
except Exception:  # pragma: no cover
    ZoneInfo = None  # type: ignore
    available_timezones = None  # type: ignore

from dotenv import load_dotenv
from flask import Flask, Response, flash, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_session import Session
from werkzeug.security import check_password_hash, generate_password_hash

_APP_DIR = os.path.dirname(os.path.abspath(__file__))
# Load .env from the project folder (next to app.py), not from the shell's cwd — otherwise
# GOOGLE_OAUTH_* and other keys are missing when Flask is started from another directory.
load_dotenv(os.path.join(_APP_DIR, ".env"))

# Google OAuth local dev: oauthlib rejects http:// callback URLs unless this is set.
# In production, use HTTPS (and do NOT set OAUTHLIB_INSECURE_TRANSPORT).
if (os.environ.get("FLASK_ENV") or "").strip().lower() != "production":
    os.environ.setdefault("OAUTHLIB_INSECURE_TRANSPORT", "1")

from wordpress_module import (
    WordPressConfig,
    create_post,
    ensure_tag_ids,
    fetch_rest_post_types,
    markdown_to_wp_html,
    upload_media,
)

import storage as _storage

_storage.init_storage()


def _ensure_initial_admin_user() -> None:
    """Env-gated one-time admin bootstrap (no hardcoded credentials).

    Seeds an admin only when BOTH ``BOOTSTRAP_ADMIN_EMAIL`` and
    ``BOOTSTRAP_ADMIN_PASSWORD`` are set in the environment and the user does
    not already exist. Leaves no default credentials in source or in a fresh DB.
    """
    sm = _storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo"
    if sm not in ("mongo", "json"):
        return
    email = (os.environ.get("BOOTSTRAP_ADMIN_EMAIL") or "").strip().lower()
    password = os.environ.get("BOOTSTRAP_ADMIN_PASSWORD") or ""
    if not email or not password:
        # No bootstrap requested — never create a default admin.
        return
    try:
        existing = _storage.get_user_by_email(email)
        if existing:
            # Ensure the bootstrap email stays admin even if older data was created differently.
            try:
                if (existing.get("role") or "").strip().lower() != "admin" and hasattr(_storage, "update_user_fields"):
                    _storage.update_user_fields(existing["id"], {"role": "admin"})
            except Exception:
                pass
            return
        _storage.insert_user(
            {
                "id": str(uuid.uuid4()),
                "email": email,
                "password_hash": generate_password_hash(password),
                "role": "admin",
                "created_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )
        app.logger.info("Created bootstrap admin user: %s", email)
    except Exception as e:
        app.logger.warning("Could not ensure bootstrap admin user: %s", e)


def _migrate_json_if_db_empty() -> None:
    """Optional one-time import from legacy JSON files into MongoDB.

    Runs only when AUTO_IMPORT_JSON=1 and the database has no projects.
    Normal operation always reads/writes MongoDB via storage only.
    """
    if (os.environ.get("AUTO_IMPORT_JSON") or "").strip() != "1":
        return
    try:
        if _storage.load_projects():
            return
        pj = os.path.join(_DATA_DIR, "projects.json")
        aj = os.path.join(_DATA_DIR, "articles.json")
        if not os.path.isfile(pj) or not os.path.isfile(aj):
            return
        with open(pj, encoding="utf-8") as f:
            pr = json.load(f)
        with open(aj, encoding="utf-8") as f:
            ar = json.load(f)
        if not isinstance(pr, list) or not isinstance(ar, list):
            return
        for p in pr:
            if isinstance(p, dict) and (p.get("id") or "").strip():
                _storage.insert_project(p)
        batch = [a for a in ar if isinstance(a, dict) and (a.get("id") or "").strip()]
        if batch:
            _storage.insert_articles_batch(batch)
    except Exception as e:
        app.logger.warning("Could not auto-import JSON into database: %s", e)


app = Flask(__name__)

# S0.5: The FastAPI service (`backend/app.main:app`) is the supported production
# backend; this legacy Flask monolith must not be served in production unless an
# operator explicitly opts in. wsgi.py imports this module, so raising here stops
# gunicorn/Procfile from booting the legacy app against a production environment.
_is_prod_env = (
    (os.environ.get("ENVIRONMENT") or "").strip().lower() == "production"
    or (os.environ.get("FLASK_ENV") or "").strip().lower() == "production"
)
_allow_legacy_flask = (os.environ.get("ALLOW_LEGACY_FLASK") or "").strip().lower() in ("1", "true", "yes", "on")
if _is_prod_env and not _allow_legacy_flask:
    raise RuntimeError(
        "Legacy Flask app.py is disabled in production. The supported backend is the FastAPI "
        "service (backend/app.main:app). Set ALLOW_LEGACY_FLASK=1 only if you intentionally run "
        "the legacy app."
    )

# Trust reverse-proxy headers (Nginx) so url_for(..., _external=True) generates HTTPS URLs.
# This is required for Google OAuth redirect_uri to match the Console configuration on production.
try:
    from werkzeug.middleware.proxy_fix import ProxyFix

    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
except Exception:
    pass

_sk = (os.environ.get("FLASK_SECRET_KEY") or "").strip()
if os.environ.get("FLASK_ENV", "").lower() == "production" and not _sk:
    raise RuntimeError("FLASK_SECRET_KEY must be set when FLASK_ENV=production")
if not _sk:
    app.secret_key = "dev-insecure-change-me"
    app.logger.warning("FLASK_SECRET_KEY not set; using insecure dev default (set FLASK_SECRET_KEY for production)")
else:
    app.secret_key = _sk

# Store sessions server-side so we can keep large generated articles.
_session_dir = os.path.join(_APP_DIR, ".flask_session")
os.makedirs(_session_dir, exist_ok=True)
app.config.update(
    SESSION_TYPE="filesystem",
    SESSION_FILE_DIR=_session_dir,
    SESSION_PERMANENT=False,
    SESSION_USE_SIGNER=True,
)
Session(app)

_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

_migrate_json_if_db_empty()
_ensure_initial_admin_user()
_ARTICLE_IMAGES_DIR = os.path.join(_DATA_DIR, "article_images")
_MAX_PROJECTS = 9  # 3×3 grid
_PRIMARY_PROTECTED_ADMIN_EMAIL = "iamakhileshsoni@gmail.com"

_orphan_owner_backfill_lock = threading.Lock()
_orphan_owner_backfill_done = False


def _norm_owner_user_id(val: Any) -> str:
    """Normalize owner id from session/Mongo (string, ObjectId, etc.)."""
    if val is None or val == "":
        return ""
    if isinstance(val, str):
        return val.strip()
    return str(val).strip()


def _maybe_backfill_orphan_project_owners() -> None:
    """Assign missing owner_user_id on legacy projects (no owner).

    Order:
    1) If ORPHAN_PROJECTS_OWNER_EMAIL is set in the environment, assign orphans to that user.
    2) Else if exactly one user exists in storage, assign orphans to that user (single-tenant).
    """
    global _orphan_owner_backfill_done
    if _orphan_owner_backfill_done:
        return
    with _orphan_owner_backfill_lock:
        if _orphan_owner_backfill_done:
            return
        try:
            if not hasattr(_storage, "list_users"):
                return
            target_uid = ""
            env_email = (os.environ.get("ORPHAN_PROJECTS_OWNER_EMAIL") or "").strip().lower()
            if env_email and hasattr(_storage, "get_user_by_email"):
                u = _storage.get_user_by_email(env_email)
                if u:
                    target_uid = _norm_owner_user_id(u.get("id"))
            if not target_uid:
                users = _storage.list_users()
                if len(users) == 1:
                    target_uid = _norm_owner_user_id(users[0].get("id"))
            if not target_uid:
                return
            for p in _storage.load_projects() or []:
                if not isinstance(p, dict):
                    continue
                if _norm_owner_user_id(p.get("owner_user_id")):
                    continue
                pid = (p.get("id") or "").strip()
                if not pid:
                    continue
                if _storage.update_project_fields(pid, {"owner_user_id": target_uid}):
                    app.logger.info("Backfilled owner_user_id on legacy project without owner: %s", pid)
        except Exception as e:
            app.logger.warning("Could not backfill orphan project owners: %s", e)
        finally:
            _orphan_owner_backfill_done = True


# Production sites linked by website URL host (see admin “Connect live projects”).
_LIVE_CONNECT_META: tuple[tuple[str, str], ...] = (
    ("Sheokand Legal", "sheokandlegal.com"),
    ("KCS", "kcsglobe.com"),
    ("TTSFM", "ttsfm.co.uk"),
)
_LIVE_CONNECT_HOSTS: frozenset[str] = frozenset(h for _, h in _LIVE_CONNECT_META)


def _normalize_url_host(raw: str) -> str:
    """Return host without www prefix, lowercase; empty if invalid."""
    s = (raw or "").strip()
    if not s:
        return ""
    try:
        if not re.match(r"^https?://", s, re.IGNORECASE):
            s = "https://" + s
        p = urlparse(s)
        host = (p.netloc or "").lower()
        if ":" in host:
            host = host.split(":")[0]
        if host.startswith("www."):
            host = host[4:]
        return host
    except Exception:
        return ""


def _project_hosts_for_match(p: dict) -> set[str]:
    hosts: set[str] = set()
    for key in ("website_url", "wp_site_url"):
        h = _normalize_url_host(p.get(key) or "")
        if h:
            hosts.add(h)
    return hosts


def _assign_live_projects_to_user_email(target_email: str) -> tuple[list[dict[str, Any]], list[str]]:
    """Set owner_user_id on projects whose URL host matches _LIVE_CONNECT_HOSTS."""
    warnings: list[str] = []
    assigned: list[dict[str, Any]] = []
    email = (target_email or "").strip().lower()
    if not email:
        return [], ["Missing email."]
    if not hasattr(_storage, "get_user_by_email"):
        return [], ["User lookup is not available."]
    u = _storage.get_user_by_email(email)
    if not u:
        return [], [f"No user found for {email}."]
    uid = _norm_owner_user_id(u.get("id"))
    if not uid:
        return [], ["User id is missing."]
    try:
        projs = _storage.load_projects()
    except Exception as e:
        return [], [f"Could not load projects: {e}"]
    found_hosts: set[str] = set()
    for p in projs or []:
        if not isinstance(p, dict):
            continue
        hosts = _project_hosts_for_match(p)
        touch = hosts & _LIVE_CONNECT_HOSTS
        if not touch:
            continue
        pid = (p.get("id") or "").strip()
        if not pid:
            continue
        if _storage.update_project_fields(pid, {"owner_user_id": uid}):
            found_hosts.update(touch)
            assigned.append(
                {
                    "id": pid,
                    "name": (p.get("name") or "").strip() or "Untitled",
                    "hosts": sorted(touch),
                }
            )
    missing = sorted(_LIVE_CONNECT_HOSTS - found_hosts)
    if missing:
        warnings.append(
            "No project found matching these host(s) (add or fix website URL on a project): "
            + ", ".join(missing)
        )
    return assigned, warnings


def _is_authenticated() -> bool:
    return bool((session.get("user_id") or "").strip())


def _current_user() -> dict | None:
    uid = (session.get("user_id") or "").strip()
    if not uid:
        return None
    return _storage.get_user_by_id(uid)


def _utc_now_str() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")


def _all_timezones() -> list[str]:
    if not available_timezones:
        return ["UTC"]
    try:
        out = sorted(list(available_timezones()))
    except Exception:
        out = ["UTC"]
    return out or ["UTC"]


def _user_timezone_name(user: dict | None) -> str:
    u = user or {}
    tz = (u.get("timezone") or "").strip()
    if tz and ZoneInfo:
        try:
            ZoneInfo(tz)
            return tz
        except Exception:
            pass
    return "UTC"


def _local_to_utc_str(dt_local: datetime, tz_name: str) -> str:
    """
    Convert a naive local datetime (picked in user timezone) to UTC schedule string.
    """
    if not ZoneInfo:
        return dt_local.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    try:
        tz = ZoneInfo(tz_name)
        aware = dt_local.replace(tzinfo=tz)
        utc = aware.astimezone(ZoneInfo("UTC"))
        return utc.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return dt_local.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def _load_plans_safe() -> dict[str, Any]:
    try:
        return _storage.load_plans() if hasattr(_storage, "load_plans") else {}
    except Exception:
        return {}


def _plan_for_user(user: dict | None) -> dict[str, Any]:
    u = user or {}
    plans = _load_plans_safe()
    key = (u.get("subscription_type") or "beta").strip().lower() or "beta"
    plan = plans.get(key) if isinstance(plans, dict) else None
    if isinstance(plan, dict) and plan:
        return plan
    fallback = plans.get("beta") if isinstance(plans, dict) else None
    return fallback if isinstance(fallback, dict) else {}


def _plan_int(plan: dict[str, Any], key: str, default: int) -> int:
    try:
        v = int(plan.get(key, default))
        return v if v >= 0 else default
    except Exception:
        return default


def _plan_bool(plan: dict[str, Any], key: str, default: bool) -> bool:
    v = plan.get(key, default)
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    s = str(v).strip().lower()
    if s in ("1", "true", "yes", "on"):
        return True
    if s in ("0", "false", "no", "off"):
        return False
    return default


def _plan_limit_summary(user: dict | None) -> dict[str, Any]:
    plan = _plan_for_user(user)
    return {
        "name": (plan.get("name") or "").strip() or "Plan",
        "max_projects": _plan_int(plan, "max_projects", 9999),
        # 0 means unlimited for the period.
        "max_articles_per_day": _plan_int(plan, "max_articles_per_day", 0),
        "max_articles_per_month": _plan_int(plan, "max_articles_per_month", 0),
        "max_writing_prompts": _plan_int(plan, "max_writing_prompts", 9999),
        "writing_prompt_char_limit": _plan_int(plan, "writing_prompt_char_limit", 100_000),
        "max_image_prompts": _plan_int(plan, "max_image_prompts", 9999),
        "image_prompt_char_limit": _plan_int(plan, "image_prompt_char_limit", 100_000),
        "allow_scheduling": _plan_bool(plan, "allow_scheduling", True),
        # 0 means unlimited per month.
        "max_scheduled_articles_per_month": _plan_int(plan, "max_scheduled_articles_per_month", 0),
        "allow_export": _plan_bool(plan, "allow_export", True),
        "allow_bulk_upload": _plan_bool(plan, "allow_bulk_upload", True),
    }


def _today_key_utc() -> str:
    return datetime.utcnow().strftime("%Y-%m-%d")


def _month_key_utc() -> str:
    return datetime.utcnow().strftime("%Y-%m")


def _int0(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except Exception:
        return default


def _article_quota_try_consume(user_id: str, add_count: int) -> tuple[bool, str]:
    """Return (ok, error_message). If ok, persists updated counters for day/month."""
    uid = (user_id or "").strip()
    n = int(add_count or 0)
    if not uid or n <= 0:
        return True, ""
    u = _storage.get_user_by_id(uid) or {}
    limits = _plan_limit_summary(u)
    day_limit = int(limits.get("max_articles_per_day") or 0)
    month_limit = int(limits.get("max_articles_per_month") or 0)
    if day_limit <= 0 and month_limit <= 0:
        return True, ""

    today = _today_key_utc()
    this_month = _month_key_utc()
    day_key = (u.get("usage_daily_articles_date") or "").strip()
    day_used = _int0(u.get("usage_daily_articles_count"), 0)
    month_key = (u.get("usage_monthly_articles_month") or "").strip()
    month_used = _int0(u.get("usage_monthly_articles_count"), 0)

    if day_key != today:
        day_key = today
        day_used = 0
    if month_key != this_month:
        month_key = this_month
        month_used = 0

    if day_limit > 0 and (day_used + n) > day_limit:
        return False, f"Daily article limit reached ({day_limit}/day). Try again tomorrow."
    if month_limit > 0 and (month_used + n) > month_limit:
        return False, f"Monthly article limit reached ({month_limit}/month). Try again next month."

    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(
                uid,
                {
                    "usage_daily_articles_date": day_key,
                    "usage_daily_articles_count": day_used + n,
                    "usage_monthly_articles_month": month_key,
                    "usage_monthly_articles_count": month_used + n,
                },
            )
    except Exception:
        # If we can't persist usage, fail closed so limits are respected.
        return False, "Could not update usage counters (database unavailable)."
    return True, ""


def _schedule_quota_try_consume(user_id: str, add_count: int) -> tuple[bool, str]:
    """
    Return (ok, error_message). If ok, persists updated scheduling counter for the current month (UTC).
    This limits how many articles a user can schedule in the current plan period (monthly renewal).
    """
    uid = (user_id or "").strip()
    n = int(add_count or 0)
    if not uid or n <= 0:
        return True, ""
    u = _storage.get_user_by_id(uid) or {}
    limits = _plan_limit_summary(u)
    month_limit = int(limits.get("max_scheduled_articles_per_month") or 0)
    if month_limit <= 0:
        return True, ""

    this_month = _month_key_utc()
    month_key = (u.get("usage_monthly_scheduled_month") or "").strip()
    month_used = _int0(u.get("usage_monthly_scheduled_count"), 0)
    if month_key != this_month:
        month_key = this_month
        month_used = 0

    if (month_used + n) > month_limit:
        return (
            False,
            f"Monthly scheduling limit reached ({month_limit}/month). Upgrade plan to schedule more articles this month.",
        )

    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(
                uid,
                {
                    "usage_monthly_scheduled_month": month_key,
                    "usage_monthly_scheduled_count": month_used + n,
                },
            )
    except Exception:
        return False, "Could not update scheduling counters (database unavailable)."

    return True, ""

def _project_ids_for_owner(user_id: str) -> list[str]:
    uid = _norm_owner_user_id(user_id)
    if not uid:
        return []
    # Fast path for Mongo: avoid loading all projects.
    try:
        if hasattr(_storage, "project_ids_for_owner") and (
            (_storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo") == "mongo"
        ):
            rows = _storage.project_ids_for_owner(uid)  # type: ignore[attr-defined]
            return [str(x).strip() for x in (rows or []) if str(x).strip()]
    except Exception:
        pass
    try:
        projs = _storage.load_projects()
    except Exception:
        return []
    sm = _storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo"
    out: list[str] = []
    for p in (projs or []):
        if not isinstance(p, dict):
            continue
        owner = _norm_owner_user_id(p.get("owner_user_id"))
        # Back-compat for legacy JSON snapshots that didn’t store owner_user_id.
        if owner and owner != uid:
            continue
        if not owner and sm == "mongo":
            continue
        pid = (p.get("id") or "").strip()
        if pid:
            out.append(pid)
    return out


def _purge_user_projects_and_files(user_id: str) -> None:
    """Remove all projects and articles for user from storage and delete local article image files."""
    uid = (user_id or "").strip()
    if not uid:
        return
    for pid in list(_project_ids_for_owner(uid)):
        try:
            for a in _articles_for_project(pid):
                aid = (a.get("id") or "").strip()
                if aid:
                    _delete_article_featured_image_file(aid)
        except Exception as e:
            app.logger.warning("Article image cleanup failed (project %s): %s", pid, e)
        if not _storage.delete_project_and_articles(pid):
            app.logger.warning("delete_project_and_articles returned false for project %s", pid)


def _articles_counts_for_owner_project_ids(project_ids: list[str]) -> dict[str, int]:
    pid_set = {x for x in (project_ids or []) if (x or "").strip()}
    if not pid_set:
        return {"total": 0, "pending": 0, "draft": 0, "published": 0, "active": 0}
    # Fast path for Mongo: group counts in the DB (no full collection scan).
    try:
        if hasattr(_storage, "count_articles_by_project_ids") and (
            (_storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo") == "mongo"
        ):
            return _storage.count_articles_by_project_ids(sorted(pid_set))  # type: ignore[attr-defined]
    except Exception:
        pass
    try:
        arts = _storage.load_articles()
    except Exception:
        arts = []
    pending = draft = published = total = 0
    for a in (arts or []):
        if not isinstance(a, dict):
            continue
        if (a.get("project_id") or "").strip() not in pid_set:
            continue
        total += 1
        st = (a.get("status") or "pending").strip().lower()
        if st == "published":
            published += 1
        elif st == "draft":
            draft += 1
        else:
            pending += 1
    return {
        "total": total,
        "pending": pending,
        "draft": draft,
        "published": published,
        "active": pending + draft,
    }


def _effective_article_usage_from_user(user: dict | None) -> dict[str, int]:
    """Day/month article counts with the same rollover rules as quota enforcement."""
    u = user or {}
    today = _today_key_utc()
    this_month = _month_key_utc()
    day_key = (u.get("usage_daily_articles_date") or "").strip()
    day_used = _int0(u.get("usage_daily_articles_count"), 0)
    month_key = (u.get("usage_monthly_articles_month") or "").strip()
    month_used = _int0(u.get("usage_monthly_articles_count"), 0)
    if day_key != today:
        day_used = 0
    if month_key != this_month:
        month_used = 0
    return {"day_used": day_used, "month_used": month_used}


def _plan_ui_for_project(project: dict, user: dict | None) -> dict[str, Any]:
    """UI flags for plan limits (Tools modal, side panel)."""
    limits = _plan_limit_summary(user)
    uid = ""
    if user and (user.get("id") or "").strip():
        uid = (user.get("id") or "").strip()
    else:
        uid = (session.get("user_id") or "").strip()
    pids = _project_ids_for_owner(uid)
    ac = _articles_counts_for_owner_project_ids(pids)
    usage = _effective_article_usage_from_user(user)
    max_day = int(limits.get("max_articles_per_day") or 0)
    max_month = int(limits.get("max_articles_per_month") or 0)
    max_wp = int(limits.get("max_writing_prompts") or 0)
    if max_wp <= 0:
        max_wp = 999999  # 0 = unlimited
    max_ip = int(limits.get("max_image_prompts") or 0)
    if max_ip <= 0:
        max_ip = 999999

    total_articles = int(ac.get("total") or 0)
    wp_count = len((project or {}).get("prompts") or [])
    ip_count = len((project or {}).get("image_prompts") or [])

    day_cap = max_day > 0 and usage["day_used"] >= max_day
    month_cap = max_month > 0 and usage["month_used"] >= max_month
    # "Max articles" (total cap) removed: only day/month generation limits apply.
    cannot_add_article = day_cap or month_cap

    allow_bulk = bool(limits.get("allow_bulk_upload", True))
    # Bulk upload is controlled only by the feature flag (not generation quota).
    lock_bulk = not allow_bulk

    lock_writing = wp_count >= max_wp
    lock_image = ip_count >= max_ip

    role = (session.get("role") or "").strip().lower()
    is_admin = role == "admin"
    upgrade_href = url_for("home", section="limits") if is_admin else url_for("home")

    reason_bulk = ""
    if not allow_bulk:
        reason_bulk = "Upgrade plan to enable Bulk Upload."

    return {
        "limits": limits,
        "usage": usage,
        "article_counts": ac,
        "lock_bulk_panel": lock_bulk,
        "lock_writing_panel": lock_writing,
        "lock_image_panel": lock_image,
        "lock_search_console_panel": False,
        "allow_scheduling": bool(limits.get("allow_scheduling", True)),
        "allow_export": bool(limits.get("allow_export", True)),
        "cannot_add_article": cannot_add_article,
        "upgrade_href": upgrade_href,
        "is_admin": is_admin,
        "reason_bulk": reason_bulk,
        "reason_writing": f"Writing prompt limit reached ({wp_count}/{max_wp})." if lock_writing else "",
        "reason_image": f"Image prompt limit reached ({ip_count}/{max_ip})." if lock_image else "",
    }


def _admin_user_profile_snapshot(u: dict[str, Any]) -> dict[str, Any]:
    uid = _norm_owner_user_id(u.get("id"))
    project_ids = _project_ids_for_owner(uid)
    articles = _articles_counts_for_owner_project_ids(project_ids)
    writing_prompts: list[dict[str, Any]] = []
    image_prompts: list[dict[str, Any]] = []
    user_projects: list[dict[str, Any]] = []
    try:
        projs = _storage.load_projects()
    except Exception:
        projs = []
    sm = _storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo"
    for p in (projs or []):
        if not isinstance(p, dict):
            continue
        owner = _norm_owner_user_id(p.get("owner_user_id"))
        if owner and owner != uid:
            continue
        if not owner and sm == "mongo":
            continue
        pid = (p.get("id") or "").strip()
        if pid:
            user_projects.append(
                {
                    "id": pid,
                    "name": (p.get("name") or "").strip() or "Untitled project",
                    "website_url": (p.get("website_url") or "").strip(),
                }
            )
        pname = (p.get("name") or "").strip() or "Untitled project"
        for pr in (p.get("prompts") or []):
            if not isinstance(pr, dict):
                continue
            writing_prompts.append(
                {
                    "project": pname,
                    "name": (pr.get("name") or "").strip() or "—",
                    "chars": len((pr.get("text") or "")),
                }
            )
        for pr in (p.get("image_prompts") or []):
            if not isinstance(pr, dict):
                continue
            image_prompts.append(
                {
                    "project": pname,
                    "name": (pr.get("name") or "").strip() or "—",
                    "chars": len((pr.get("text") or "")),
                }
            )
    try:
        user_projects.sort(key=lambda x: ((x.get("name") or "").strip().lower(), (x.get("id") or "")))
    except Exception:
        pass
    return {
        **u,
        "stats_projects_total": len(project_ids),
        "stats_projects": user_projects,
        "stats_articles": articles,
        "stats_writing_prompts": writing_prompts,
        "stats_image_prompts": image_prompts,
    }


@app.before_request
def _touch_last_activity():
    if not _is_authenticated():
        return None
    uid = (session.get("user_id") or "").strip()
    if not uid:
        return None
    last = session.get("_last_activity_touch") or ""
    now_ts = int(time.time())
    try:
        last_ts = int(last) if str(last).isdigit() else 0
    except Exception:
        last_ts = 0
    # Avoid writing to storage on every request.
    if now_ts - last_ts < 300:
        return None
    session["_last_activity_touch"] = str(now_ts)
    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(uid, {"last_activity_at": _utc_now_str()})
    except Exception:
        pass
    return None


@app.before_request
def _require_login_for_app() -> None:
    # Allow auth endpoints and static assets without login.
    ep = request.endpoint or ""
    if ep in {"home", "auth_login", "auth_register", "auth_logout", "static"}:
        return
    # Let Flask handle None endpoints (404 etc).
    if not ep:
        return
    if not _is_authenticated():
        return redirect(url_for("home"))


def _ensure_data_dir() -> None:
    os.makedirs(_DATA_DIR, exist_ok=True)


def _load_projects() -> list[dict]:
    """Projects visible in the workspace list (current user only, including admins)."""
    projects = _storage.load_projects()
    user_id = _norm_owner_user_id(session.get("user_id"))
    if not user_id:
        return []
    out: list[dict] = []
    for p in projects:
        if not isinstance(p, dict):
            continue
        if _norm_owner_user_id(p.get("owner_user_id")) == user_id:
            out.append(p)
    return out


def _update_project_fields(project_id: str, updates: dict) -> bool:
    return _storage.update_project_fields(project_id, updates)


_wp_scheduled_processor_lock = threading.Lock()
_wp_bg_trigger_last = 0.0
_wp_bg_trigger_lock = threading.Lock()


def _load_articles() -> list[dict]:
    data = _storage.load_articles()
    for a in data:
        if not (a.get("status") or "").strip():
            a["status"] = "pending"
        gs = (a.get("gsc_status") or "").strip().lower()
        # Back-compat: older versions stored "requested" for a successful URL Inspection call.
        if gs == "requested":
            gs = "inspected"
            a["gsc_status"] = "inspected"
        if gs not in {"pending", "inspected"}:
            a["gsc_status"] = "pending"
    return data


def _get_project_by_id(project_id: str) -> dict | None:
    pid = (project_id or "").strip()
    if not pid:
        return None
    p = _get_project_by_id_unscoped(pid)
    if not p:
        return None
    user_id = _norm_owner_user_id(session.get("user_id"))
    role = (session.get("role") or "").strip().lower()
    owner = _norm_owner_user_id(p.get("owner_user_id"))
    if owner == user_id:
        return p
    if role == "admin":
        return p
    return None


def _get_project_by_id_unscoped(project_id: str) -> dict | None:
    """Load a project from storage without session/owner filtering.

    Used by background threads and APScheduler where no request context exists.
    HTTP routes must enforce access before calling code that relies on this.
    """
    pid = (project_id or "").strip()
    for p in _storage.load_projects():
        if (p.get("id") or "") == pid:
            return p
    return None


_bulk_schedule_error_lock = threading.Lock()
_bulk_schedule_last_errors: dict[str, str] = {}

def _llm_timeout_seconds() -> int:
    try:
        v = int(os.environ.get("LLM_TIMEOUT_SECONDS", "120") or 120)
        return max(20, min(v, 600))
    except Exception:
        return 120


def _retry_sleep_seconds(attempt: int) -> float:
    # Exponential-ish backoff with jitter, capped.
    base = min(12.0, 1.5 * (2 ** max(0, attempt - 1)))
    return base + random.random() * 0.6


def _with_retries(fn, *, tries: int = 3, label: str = "operation"):
    last_err: Exception | None = None
    for i in range(1, tries + 1):
        try:
            return fn()
        except Exception as e:
            last_err = e
            if i >= tries:
                break
            time.sleep(_retry_sleep_seconds(i))
    raise last_err or RuntimeError(f"{label} failed")


def _set_bulk_schedule_error(project_id: str, message: str) -> None:
    pid = (project_id or "").strip()
    if not pid:
        return
    with _bulk_schedule_error_lock:
        _bulk_schedule_last_errors[pid] = message


def _pop_bulk_schedule_error(project_id: str) -> str | None:
    pid = (project_id or "").strip()
    if not pid:
        return None
    with _bulk_schedule_error_lock:
        return _bulk_schedule_last_errors.pop(pid, None)


def _normalize_project_prompts(project: dict) -> None:
    """Ensure prompts list exists and entries have id, name, text."""
    raw = project.get("prompts")
    if not isinstance(raw, list):
        project["prompts"] = []
        return
    cleaned: list[dict] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        pid = (item.get("id") or "").strip()
        name = (item.get("name") or "").strip()
        text = (item.get("text") or "").strip()
        if not pid or not name:
            continue
        cleaned.append({"id": pid, "name": name[:200], "text": text[:100_000]})
    project["prompts"] = cleaned


def _get_prompt_by_id(project: dict, prompt_id: str) -> dict | None:
    pid = (prompt_id or "").strip()
    for pr in project.get("prompts") or []:
        if isinstance(pr, dict) and (pr.get("id") or "") == pid:
            return pr
    return None


def _resolve_default_prompt_text(project: dict) -> str | None:
    """Text of the default prompt for article generation, or None to use built-in template."""
    _normalize_project_prompts(project)
    did = (project.get("default_prompt_id") or "").strip()
    if not did:
        return None
    pr = _get_prompt_by_id(project, did)
    if pr and (pr.get("text") or "").strip():
        return (pr.get("text") or "").strip()
    return None


# Placeholders users can put in stored prompt text; filled at generation time.
_ARTICLE_PROMPT_PLACEHOLDER_KEYS = (
    "{article title}",
    "{targeting keywords}",
    "{focus keyphrase}",
    "{title}",
    "{keywords}",
    "{focus_keyphrase}",
)


def _prompt_template_has_placeholders(raw: str) -> bool:
    s = raw or ""
    return any(k in s for k in _ARTICLE_PROMPT_PLACEHOLDER_KEYS)


def _interpolate_article_prompt_template(
    raw: str,
    title: str,
    keywords: list[str],
    focus_keyphrase: str | None = None,
) -> str:
    """Replace {article title}, {targeting keywords}, {focus keyphrase}, and short aliases."""
    kw_display = ", ".join(keywords) if keywords else "(none)"
    fk_display = (focus_keyphrase or "").strip() or "(none)"
    out = raw or ""
    # Longer tokens first so partial overlaps are safe
    out = out.replace("{article title}", title)
    out = out.replace("{targeting keywords}", kw_display)
    out = out.replace("{focus keyphrase}", fk_display)
    out = out.replace("{title}", title)
    out = out.replace("{keywords}", kw_display)
    out = out.replace("{focus_keyphrase}", fk_display)
    return out


def _normalize_project_image_settings(project: dict) -> None:
    """Per-project featured image style + ChatGPT prompt optimizer toggle."""
    if not isinstance(project, dict):
        return
    style = (project.get("image_style") or "").strip().lower()
    if style not in ("semi_real", "photorealistic", "illustration"):
        project["image_style"] = "semi_real"
    oip = project.get("optimize_image_prompt")
    if oip is None:
        project["optimize_image_prompt"] = True
    else:
        project["optimize_image_prompt"] = bool(oip)


def _normalize_project_image_prompts(project: dict) -> None:
    """Ensure image_prompts list exists (same shape as writing prompts)."""
    raw = project.get("image_prompts")
    if not isinstance(raw, list):
        project["image_prompts"] = []
        _normalize_project_image_settings(project)
        return
    cleaned: list[dict] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        pid = (item.get("id") or "").strip()
        name = (item.get("name") or "").strip()
        text = (item.get("text") or "").strip()
        if not pid or not name:
            continue
        cleaned.append({"id": pid, "name": name[:200], "text": text[:100_000]})
    project["image_prompts"] = cleaned
    _normalize_project_image_settings(project)


def _normalize_project_gsc(project: dict) -> None:
    """Search Console property URL + whether to request inspection after a live WordPress publish."""
    raw = project.get("gsc_property_url")
    project["gsc_property_url"] = (raw or "").strip() if isinstance(raw, str) else ""
    iop = project.get("gsc_index_on_publish")
    if iop is None:
        project["gsc_index_on_publish"] = True
    else:
        project["gsc_index_on_publish"] = bool(iop)


def _normalize_project_context_links(project: dict) -> None:
    """Ensure context_links list exists: phrase + url for inline linking when posting to WordPress."""
    raw = project.get("context_links")
    if not isinstance(raw, list):
        project["context_links"] = []
        return
    cleaned: list[dict] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        lid = (item.get("id") or "").strip()
        phrase = (item.get("phrase") or item.get("text") or "").strip()
        url = (item.get("url") or item.get("link") or "").strip()
        if not lid or not phrase:
            continue
        cleaned.append({"id": lid, "phrase": phrase[:2000], "url": url[:2048]})
    project["context_links"] = cleaned


def _normalize_project_wp_defaults(project: dict) -> None:
    """
    Ensure default WordPress settings exist on the project.
    - default_wp_rest_base: REST collection (e.g. posts, pages, articles)
    - default_wp_status: draft or publish
    """
    if not isinstance(project, dict):
        return
    rb = (project.get("default_wp_rest_base") or "").strip().strip("/") or "posts"
    project["default_wp_rest_base"] = rb
    st = (project.get("default_wp_status") or "draft").strip().lower()
    project["default_wp_status"] = st if st in ("draft", "publish") else "draft"


def _is_valid_context_link_url(url: str) -> bool:
    raw = (url or "").strip()
    if not raw or ")" in raw:
        return False
    u = urlparse(raw)
    if u.scheme not in ("http", "https") or not u.netloc:
        return False
    return True


def _context_link_rules_for_project(project: dict) -> list[tuple[str, str]]:
    """(phrase, url) pairs, longest phrases first; skips invalid rows."""
    _normalize_project_context_links(project)
    rules: list[tuple[str, str]] = []
    for item in project.get("context_links") or []:
        phrase = (item.get("phrase") or "").strip()
        url = (item.get("url") or "").strip()
        if not phrase or not url or not _is_valid_context_link_url(url):
            continue
        if "]" in phrase or "\n" in phrase or "\r" in phrase:
            continue
        rules.append((phrase, url))
    rules.sort(key=lambda x: len(x[0]), reverse=True)
    return rules


_CTX_LINK_TOKEN = "\u2060CTX\u2060{}"

def _protect_regions_for_context_links(md: str) -> tuple[str, list[str]]:
    """Temporarily replace markdown links and HTML anchors so phrase replacement does not run inside them."""
    vault: list[str] = []

    def stash_md(m):
        vault.append(m.group(0))
        return _CTX_LINK_TOKEN.format(len(vault) - 1)

    s = re.sub(r"\[[^\]]*\]\([^)]*\)", stash_md, md)

    def stash_html(m):
        vault.append(m.group(0))
        return _CTX_LINK_TOKEN.format(len(vault) - 1)

    s = re.sub(r"<a\s[^>]*>.*?</a>", stash_html, s, flags=re.I | re.DOTALL)
    return s, vault


def _unprotect_context_link_regions(s: str, vault: list[str]) -> str:
    for i, chunk in enumerate(vault):
        s = s.replace(_CTX_LINK_TOKEN.format(i), chunk)
    return s


def _apply_context_links_to_markdown(md: str, project: dict | None) -> str:
    """
    Wrap phrase matches in markdown links before HTML conversion.
    Matching is case-insensitive; the linked text keeps the article’s original casing.
    Longer phrases are applied first; existing markdown links and <a> tags are not modified.
    """
    if not md or not project:
        return md
    rules = _context_link_rules_for_project(project)
    if not rules:
        return md
    s = md
    for phrase, url in rules:
        s, vault = _protect_regions_for_context_links(s)
        pattern = re.compile(re.escape(phrase), re.IGNORECASE)
        if not pattern.search(s):
            s = _unprotect_context_link_regions(s, vault)
            continue

        def _ctx_link_repl(m):
            matched = m.group(0)
            return f"[{matched}]({url})"

        s = pattern.sub(_ctx_link_repl, s)
        s = _unprotect_context_link_regions(s, vault)
    return s


def _article_body_to_wp_html(body: str, project: dict | None) -> str:
    """Markdown → HTML for WordPress, applying project context links when `project` is set."""
    md = _apply_context_links_to_markdown(body or "", project)
    return markdown_to_wp_html(md)


def _project_wp_credentials_configured(project: dict) -> bool:
    """True when WordPress username and application password are set (required for posting)."""
    u = (project.get("wp_username") or "").strip()
    p = (project.get("wp_app_password") or "").strip()
    return bool(u and p)


def _get_image_prompt_by_id(project: dict, prompt_id: str) -> dict | None:
    pid = (prompt_id or "").strip()
    for pr in project.get("image_prompts") or []:
        if isinstance(pr, dict) and (pr.get("id") or "") == pid:
            return pr
    return None


def _resolve_default_image_prompt_text(project: dict) -> str | None:
    _normalize_project_image_prompts(project)
    did = (project.get("default_image_prompt_id") or "").strip()
    if not did:
        return None
    pr = _get_image_prompt_by_id(project, did)
    if pr and (pr.get("text") or "").strip():
        return (pr.get("text") or "").strip()
    return None


def _ensure_article_images_dir() -> None:
    os.makedirs(_ARTICLE_IMAGES_DIR, exist_ok=True)


def _article_featured_image_path(article_id: str) -> str:
    return os.path.join(_ARTICLE_IMAGES_DIR, f"{article_id}.png")


def _delete_article_featured_image_file(article_id: str) -> None:
    p = _article_featured_image_path(article_id)
    if os.path.isfile(p):
        try:
            os.remove(p)
        except OSError:
            pass


def _save_article_featured_image_png(article_id: str, png_bytes: bytes) -> None:
    _ensure_article_images_dir()
    path = _article_featured_image_path(article_id)
    with open(path, "wb") as f:
        f.write(png_bytes)


_MAX_FEATURED_UPLOAD_BYTES = 8 * 1024 * 1024


_PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"


def _convert_upload_to_png_bytes(raw: bytes) -> bytes:
    """Convert a single uploaded image (JPEG/PNG/WebP/GIF) to PNG bytes."""
    from io import BytesIO

    try:
        from PIL import Image
    except ImportError:
        # Allow PNG-only uploads when Pillow is not installed (same bytes stored as .png).
        if len(raw) >= 8 and raw[:8] == _PNG_SIGNATURE:
            return raw
        raise ValueError(
            "Install Pillow to upload JPEG, WebP, or GIF: pip install Pillow "
            "(PNG uploads work without it; see requirements.txt)."
        ) from None

    im = Image.open(BytesIO(raw))
    if im.mode == "P" and "transparency" in im.info:
        im = im.convert("RGBA")
    if im.mode == "LA":
        im = im.convert("RGBA")
    if im.mode == "RGBA":
        bg = Image.new("RGB", im.size, (255, 255, 255))
        bg.paste(im, mask=im.split()[3])
        im = bg
    elif im.mode != "RGB":
        im = im.convert("RGB")
    out = BytesIO()
    im.save(out, format="PNG", optimize=True)
    return out.getvalue()


_FEATURED_IMAGE_META_KEYS = (
    "featured_image_prompt_final",
    "featured_image_prompt_raw",
    "featured_image_model",
    "featured_image_quality",
    "featured_image_size",
    "featured_image_optimizer_model",
    "featured_image_prompt_optimizer_error",
)

_IMAGE_STYLE_INSTRUCTIONS: dict[str, str] = {
    "semi_real": (
        "Semi-realistic: cinematic editorial hero image, slightly stylized 3D or painterly realism, "
        "dramatic but believable lighting, high detail, NOT a flat cartoon, NOT corporate clipart."
    ),
    "photorealistic": (
        "Photorealistic: natural photography look, realistic lens and depth of field, natural lighting and materials."
    ),
    "illustration": (
        "Illustration: clean vector or flat illustration suitable for web, bold shapes, limited palette, "
        "not photorealistic."
    ),
}


def _clear_featured_image_generation_meta_updates() -> dict[str, str]:
    """Clear stored OpenAI image / prompt metadata on an article."""
    return {k: "" for k in _FEATURED_IMAGE_META_KEYS}


def _default_openai_image_size_for_model(model: str) -> str:
    """
    DALL·E 3 supports wide sizes like 1792x1024; GPT Image models only allow
    1024x1024, 1024x1536, 1536x1024, and auto (per API error messages).
    """
    m = (model or "").lower()
    if "dall-e" in m:
        return "1792x1024"
    return "1536x1024"


def _coerce_image_size_for_model(model: str, size: str) -> str:
    """Map legacy DALL·E sizes to the closest GPT Image size when needed."""
    m = (model or "").lower()
    s = (size or "").strip().lower()
    if not s:
        return _default_openai_image_size_for_model(model)
    if "dall-e" in m:
        return s
    gpt_allowed = {"1024x1024", "1024x1536", "1536x1024", "auto"}
    if s in gpt_allowed:
        return s
    if s == "1792x1024":
        return "1536x1024"
    if s == "1024x1792":
        return "1024x1536"
    return s


def _openai_image_env_config() -> dict[str, str]:
    """Env-driven OpenAI Images API settings (see README)."""
    model = (os.environ.get("OPENAI_IMAGE_MODEL", "gpt-image-1.5") or "gpt-image-1.5").strip()
    env_size = (os.environ.get("OPENAI_IMAGE_SIZE") or "").strip()
    size = env_size or _default_openai_image_size_for_model(model)
    return {
        "model": model,
        "quality": (os.environ.get("OPENAI_IMAGE_QUALITY", "high") or "").strip(),
        "size": size,
        "response_format": (os.environ.get("OPENAI_IMAGE_RESPONSE_FORMAT", "b64_json") or "b64_json").strip(),
    }


def _normalize_image_quality_for_model(model: str, raw_quality: str) -> str | None:
    """Map env quality to values supported by the selected image model family."""
    m = (model or "").lower()
    q = (raw_quality or "").strip().lower()
    if not q:
        return None
    if "dall-e" in m:
        if q in ("high", "hd"):
            return "hd"
        if q in ("medium", "standard", "low", "auto"):
            return "standard"
        return q if q in ("standard", "hd") else "standard"
    if q in ("low", "medium", "high", "auto"):
        return q
    if q in ("hd", "high"):
        return "high"
    if q in ("standard", "medium"):
        return "medium"
    return q


def _decode_openai_image_response(resp: Any) -> bytes:
    """Return PNG/JPEG bytes from an OpenAI images.generate response."""
    if not resp or not getattr(resp, "data", None):
        raise ValueError("Image API returned no data.")
    item = resp.data[0]
    b64 = getattr(item, "b64_json", None)
    if b64:
        return base64.b64decode(b64)
    url = getattr(item, "url", None)
    if url:
        with urlopen(url, timeout=120) as r:
            return r.read()
    raise ValueError("Image API returned neither b64_json nor url.")


def _generate_featured_image_png_bytes(image_prompt: str) -> tuple[bytes, dict[str, Any]]:
    """
    OpenAI Images API (model from OPENAI_IMAGE_MODEL, default gpt-image-1.5).
    Requires OPENAI_API_KEY. Returns (png_bytes, metadata for article record).
    """
    from openai import OpenAI

    key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not key:
        raise ValueError(
            "OPENAI_API_KEY is required for featured image generation. "
            "Groq and other chat-only keys cannot generate images."
        )
    client = OpenAI(api_key=key)
    prompt = (image_prompt or "").strip()[:4000]
    if not prompt:
        raise ValueError("Image prompt is empty.")

    cfg = _openai_image_env_config()
    model = cfg["model"]
    size = _coerce_image_size_for_model(model, cfg["size"])
    raw_q = cfg["quality"]
    rf = cfg["response_format"].lower()
    norm_q = _normalize_image_quality_for_model(model, raw_q)

    attempts: list[dict[str, Any]] = []
    base: dict[str, Any] = {"model": model, "prompt": prompt, "n": 1}
    if size:
        base["size"] = size
    if norm_q:
        base["quality"] = norm_q

    if rf and rf not in ("none", "auto", ""):
        attempts.append({**base, "response_format": rf})
    attempts.append(dict(base))

    last_err: Exception | None = None
    for kwargs in attempts:
        try:
            def _call():
                # openai-python supports request timeouts via kwargs; if unsupported, this is ignored.
                return client.images.generate(**{**kwargs, "timeout": _llm_timeout_seconds()})

            resp = _with_retries(_call, tries=3, label="image generation")
            png_bytes = _decode_openai_image_response(resp)
            meta: dict[str, Any] = {
                "featured_image_model": model,
                "featured_image_quality": str(kwargs.get("quality") or ""),
                "featured_image_size": size,
                "featured_image_prompt_final": prompt,
            }
            return png_bytes, meta
        except Exception as e:
            last_err = e
            continue
    msg = str(last_err) if last_err else "unknown error"
    raise ValueError(f"Image generation failed: {msg}") from last_err


def _articles_for_project(project_id: str) -> list[dict]:
    pid = (project_id or "").strip()
    return [a for a in _load_articles() if (a.get("project_id") or "") == pid]


def _get_article_by_id(article_id: str) -> dict | None:
    aid = (article_id or "").strip()
    for a in _load_articles():
        if (a.get("id") or "") == aid:
            return a
    return None


def _update_article_fields(article_id: str, updates: dict) -> bool:
    return _storage.update_article_fields(article_id, updates)


def _article_wp_scheduled_at_str(article: dict | None) -> str:
    """Normalize schedule time from article row (string or BSON datetime from MongoDB)."""
    if not article:
        return ""
    return _storage._coerce_wp_scheduled_at_str(article.get("wp_scheduled_at"))


def _article_wp_scheduled_at_utc_str(article: dict | None) -> str:
    """UTC schedule timestamp (preferred for due checks)."""
    if not article:
        return ""
    v = article.get("wp_scheduled_at_utc")
    if isinstance(v, str) and v.strip():
        return v.strip()
    # Back-compat: derive UTC if we have local time + tz offset.
    local_str = _storage._coerce_wp_scheduled_at_str(article.get("wp_scheduled_at"))
    if not local_str:
        return ""
    off = article.get("wp_schedule_tz_offset_min")
    try:
        off_i = int(off)
    except Exception:
        return ""
    try:
        dt_local = datetime.strptime(local_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return ""
    dt_utc = dt_local + timedelta(minutes=off_i)
    return dt_utc.replace(microsecond=0).strftime("%Y-%m-%d %H:%M:%S")


def _app_article_status_from_wp_rest_status(raw) -> str:
    """Map WordPress REST post `status` to our article status (published vs draft)."""
    if isinstance(raw, str) and raw.strip():
        w = raw.strip().lower()
    else:
        w = str(raw or "draft").lower()
    return "published" if w == "publish" else "draft"


def _article_to_last(article: dict) -> dict:
    """Shape compatible with article generator / WordPress templates."""
    return {
        "title": article.get("title") or "",
        "keywords": article.get("keywords") or [],
        "article": article.get("article") or "",
        "focus_keyphrase": article.get("focus_keyphrase") or "",
        "meta_title": article.get("meta_title") or "",
        "meta_description": article.get("meta_description") or "",
        "generated_at": article.get("generated_at") or "",
    }


def _normalize_article_text(s: str | None) -> str:
    if not s:
        return ""
    return s.replace("\r\n", "\n").replace("\r", "\n")


def _article_body_matches_stored(form_body: str | None, stored: str | None) -> bool:
    return _normalize_article_text(form_body) == _normalize_article_text(stored)


def _parse_article_datetime(s: str) -> datetime | None:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _parse_schedule_times_json(raw: str) -> dict[str, str] | None:
    """Parse JSON object of article_id -> datetime-local string from bulk schedule form."""
    s = (raw or "").strip()
    if not s:
        return None
    try:
        data = json.loads(s)
    except (json.JSONDecodeError, TypeError, ValueError):
        return None
    if not isinstance(data, dict):
        return None
    out: dict[str, str] = {}
    for k, v in data.items():
        ks = str(k).strip()
        if not ks or v is None:
            continue
        vs = str(v).strip()
        if vs:
            out[ks] = vs
    return out or None


def _parse_bulk_schedule_datetime(raw: str) -> datetime | None:
    """Parse datetime from bulk schedule form (datetime-local or stored format)."""
    s = (raw or "").strip().replace("T", " ", 1)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _clamp_future_schedule_dt(dt: datetime) -> datetime:
    # Clamp using UTC semantics where possible (posting uses wp_scheduled_at_utc).
    # Here we only ensure the timestamp isn't in the past relative to server time.
    now = datetime.utcnow().replace(microsecond=0)
    if dt < now:
        return now + timedelta(minutes=1)
    return dt.replace(microsecond=0)


def _bulk_schedule_thread_entry(project_id: str, ids: list[str], form_snapshot: dict) -> None:
    try:
        with app.app_context():
            err = _bulk_schedule_pipeline(project_id, ids, form_snapshot)
            if err:
                app.logger.error("Bulk schedule failed: %s", err)
                _set_bulk_schedule_error(project_id, err)
            else:
                with _bulk_schedule_error_lock:
                    _bulk_schedule_last_errors.pop((project_id or "").strip(), None)
    except Exception as e:
        app.logger.exception("Bulk schedule raised")
        _set_bulk_schedule_error(project_id, str(e) or "Scheduling failed unexpectedly.")


def _bulk_schedule_pipeline(project_id: str, ids: list[str], form_snapshot: dict) -> str | None:
    project = _get_project_by_id_unscoped(project_id)
    if not project:
        return "Project not found."

    bulk_prompt = (form_snapshot.get("bulk_prompt_id") or "").strip()
    bulk_image = (form_snapshot.get("bulk_image_prompt_id") or "").strip()

    proj_gen = dict(project)
    _normalize_project_prompts(proj_gen)
    _normalize_project_image_prompts(proj_gen)

    arts = _load_articles()
    id_set = set(ids)
    targets = [
        a
        for a in arts
        if (a.get("id") or "") in id_set and (a.get("project_id") or "") == project_id
    ]
    if not targets:
        return "No matching articles for this project."

    proj_check = dict(project)
    _normalize_project_image_prompts(proj_check)
    need_img = len(proj_check.get("image_prompts") or []) > 0
    # Quota enforcement for bulk generation inside scheduling.
    # Bulk scheduling may generate multiple articles at once; consume quota per generated article.
    uid = (project.get("owner_user_id") or "").strip() or (session.get("user_id") or "").strip()

    for t in targets:
        aid = (t.get("id") or "").strip()
        if not aid or t.get("wp_post_id"):
            continue
        need_body = not (t.get("article") or "").strip()
        need_img_file = need_img and not os.path.isfile(_article_featured_image_path(aid))
        if not need_body and not need_img_file:
            continue
        # Mark progress for UI (scheduled items should show "Preparing" until ready).
        _update_article_fields(
            aid,
            {
                "wp_schedule_state": "preparing",
                "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            },
        )
        title = (t.get("title") or "").strip()
        if not title:
            return "Cannot generate or schedule: one or more selected articles have no title. Add a title first."
        try:
            ok_quota, quota_err = _article_quota_try_consume(uid, 1)
            if not ok_quota:
                raise ValueError(quota_err or "Current limit is exhausted. Upgrade plan to continue.")

            cur_attempts = _int0((_get_article_by_id(aid) or {}).get("wp_generation_attempts"), 0)
            _update_article_fields(aid, {"wp_generation_attempts": cur_attempts + 1})

            kws = _article_keywords_list(t)
            ok, err, img_err = _generate_article_content_core(
                proj_gen,
                aid,
                title=title,
                keywords=kws,
                writing_prompt_id=bulk_prompt,
                image_prompt_id=bulk_image,
                user_focus_keyphrase=((t.get("focus_keyphrase") or "").strip() or None),
            )
            if not ok:
                raise ValueError(err or "Generation failed.")
            if need_img and not os.path.isfile(_article_featured_image_path(aid)):
                raise ValueError(
                    "Featured image missing after generation. "
                    + (img_err or "Check OPENAI_API_KEY and OPENAI_IMAGE_* settings.")
                )

            _update_article_fields(
                aid,
                {
                    "wp_schedule_state": "ready",
                    "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "wp_schedule_error": "",
                },
            )
        except Exception as e:
            # Keep schedule intact; surface error and stop background work.
            _update_article_fields(
                aid,
                {
                    "wp_schedule_state": "error",
                    "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "wp_schedule_error": str(e)[:500],
                },
            )
            return f"Generation failed for «{title[:80]}»: {e}"
    return None


def _earlier_batch_member_blocks(arts: list[dict], article: dict, now: datetime) -> bool:
    """
    True if another article in the same batch with a lower index is still ahead of this one
    in the queue: they have no wp_post_id yet, their schedule time is already due (<= now),
    and they still have a schedule row. Earlier rows with a *future* schedule time do not block
    (avoids deadlock when batch times are out of order or were edited).
    """
    bid = (article.get("wp_schedule_batch_id") or "").strip()
    if not bid:
        return False
    my_idx = int(article.get("wp_schedule_batch_index") or 0)
    for o in arts:
        if (o.get("wp_schedule_batch_id") or "").strip() != bid:
            continue
        oidx = int(o.get("wp_schedule_batch_index") or 0)
        if oidx >= my_idx:
            continue
        # Don't let a failed earlier item block the rest of the batch forever.
        if (o.get("wp_schedule_state") or "").strip().lower() == "error":
            continue
        if o.get("wp_post_id"):
            continue
        sched_o = _article_wp_scheduled_at_utc_str(o) or _article_wp_scheduled_at_str(o)
        if not sched_o:
            continue
        try:
            dt_o = datetime.strptime(sched_o, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
        if dt_o > now:
            continue
        return True
    return False


def _redirect_project_detail_with_dates(project_id: str, *, watch_schedule: bool = False):
    df = request.form.get("date_from", "").strip()
    dt = request.form.get("date_to", "").strip()
    fs = request.form.get("filter_status", "").strip().lower()
    if fs not in {"pending", "draft", "published"}:
        fs = ""
    pairs = [("date_from", df), ("date_to", dt), ("status", fs)]
    if watch_schedule:
        pairs.append(("watch_schedule", "1"))
    q = urlencode({k: v for k, v in pairs if v})
    url = url_for("project_detail", project_id=project_id)
    return redirect(url + ("?" + q if q else ""))


def _cleared_wp_schedule_fields() -> dict:
    return {
        "wp_scheduled_at": "",
        "wp_scheduled_at_utc": "",
        "wp_schedule_wp_status": "",
        "wp_schedule_error": "",
        "wp_schedule_batch_id": "",
        "wp_schedule_batch_index": "",
        "wp_schedule_batch_total": "",
        "wp_schedule_state": "",
        "wp_schedule_state_updated_at": "",
        "wp_schedule_next_retry_at": "",
        "wp_schedule_tz_offset_min": "",
    }


def _pending_scheduled_article_rows(project_id: str) -> list[dict]:
    """Articles queued for WordPress (scheduled time set, not yet posted)."""
    pid = (project_id or "").strip()
    proj = _get_project_by_id_unscoped(pid) or {}
    proj_check = dict(proj) if isinstance(proj, dict) else {}
    _normalize_project_image_prompts(proj_check)
    need_img = len(proj_check.get("image_prompts") or []) > 0
    rows: list[dict] = []
    try:
        arts = (
            _storage.load_scheduled_pending_for_project_minimal(pid, limit=300)
            if hasattr(_storage, "load_scheduled_pending_for_project_minimal")
            else _articles_for_project(pid)
        )
    except Exception:
        arts = _articles_for_project(pid)
    for a in arts:
        if a.get("wp_post_id"):
            continue
        sched = _article_wp_scheduled_at_str(a)
        if not sched:
            continue
        wp_st = (a.get("wp_schedule_wp_status") or "draft").strip().lower()
        if wp_st not in ("draft", "publish"):
            wp_st = "draft"
        aid = (a.get("id") or "").strip()
        has_body = bool((a.get("article") or "").strip())
        has_img = bool(aid and os.path.isfile(_article_featured_image_path(aid)))
        ready = bool(has_body and ((not need_img) or has_img))
        state = (a.get("wp_schedule_state") or "").strip().lower()
        if state not in {"queued", "preparing", "ready", "posting", "error"}:
            state = ""
        err = (a.get("wp_schedule_error") or "").strip()
        rows.append(
            {
                "id": aid,
                "title": (a.get("title") or "").strip() or "(Untitled)",
                "wp_scheduled_at": sched,
                "wp_schedule_wp_status": wp_st,
                "ready": ready,
                "state": state,
                "error": err,
            }
        )

    def _sort_key(r: dict):
        dt = _parse_article_datetime(r.get("wp_scheduled_at") or "")
        return dt or datetime.min

    rows.sort(key=_sort_key)
    return rows


def _bulk_set_wp_schedule_fields_now(project_id: str, ids: list[str], form_snapshot: dict) -> tuple[int, int, str | None]:
    """
    Persist schedule metadata immediately so UI can show queued items right away.
    Returns (scheduled_count, skipped_count, error_message).
    """
    project = _get_project_by_id_unscoped(project_id)
    if not project:
        return 0, 0, "Project not found."

    wp_st = (form_snapshot.get("schedule_wp_status") or "draft").strip().lower()
    if wp_st not in ("draft", "publish"):
        wp_st = "draft"

    wp_types, _ = _wp_post_types_for_project(project)
    allowed_bases = {t["rest_base"] for t in wp_types}
    raw_schedule_rb = (form_snapshot.get("schedule_wp_rest_base") or "").strip()
    if raw_schedule_rb:
        schedule_rest_base = _normalize_wp_rest_base(raw_schedule_rb, allowed_bases)
    else:
        schedule_rest_base = _normalize_wp_rest_base(project.get("default_wp_rest_base"), allowed_bases)

    times_map = _parse_schedule_times_json(form_snapshot.get("schedule_times_json") or "")
    if not times_map:
        return 0, 0, "Please set a date and time for each selected article."

    owner = _storage.get_user_by_id((project.get("owner_user_id") or "").strip()) if hasattr(_storage, "get_user_by_id") else None
    tz_name = _user_timezone_name(owner)

    arts = _load_articles()
    id_set = set(ids)
    targets = [
        a
        for a in arts
        if (a.get("id") or "") in id_set and (a.get("project_id") or "") == project_id
    ]
    if not targets:
        return 0, 0, "No matching articles for this project."

    # Stable order for batch semantics (created_at asc).
    targets.sort(key=lambda x: ((x.get("created_at") or ""), (x.get("id") or "")))

    batch_id = str(uuid.uuid4())
    bulk_updates: list[tuple[str, dict]] = []
    skipped = 0
    sched_idx = 0
    for t in targets:
        aid = (t.get("id") or "").strip()
        if not aid:
            continue
        wp_post_id = t.get("wp_post_id")
        if str(wp_post_id).strip() if wp_post_id is not None else "":
            skipped += 1
            continue
        raw_dt = times_map.get(aid)
        if raw_dt is None:
            title_hint = (t.get("title") or "").strip() or "Untitled"
            return 0, 0, f"Missing schedule time for «{title_hint[:80]}»."
        dt = _parse_bulk_schedule_datetime(raw_dt)
        if not dt:
            title_hint = (t.get("title") or "").strip() or "Untitled"
            return 0, 0, f"Invalid date and time for «{title_hint[:80]}»."
        # Stored "wp_scheduled_at" is the user-entered local timestamp (for display),
        # while "wp_scheduled_at_utc" is used for posting.
        dt_local = dt.replace(microsecond=0)
        dt_utc_str = _local_to_utc_str(dt_local, tz_name)
        try:
            dt_utc = datetime.strptime(dt_utc_str, "%Y-%m-%d %H:%M:%S")
        except ValueError:
            dt_utc = dt_local
        dt_utc = _clamp_future_schedule_dt(dt_utc)
        sched_str = dt_local.strftime("%Y-%m-%d %H:%M:%S")
        sched_utc_str = dt_utc.strftime("%Y-%m-%d %H:%M:%S")
        bulk_updates.append(
            (
                aid,
                {
                    "wp_scheduled_at": sched_str,
                    "wp_scheduled_at_utc": sched_utc_str,
                    "wp_schedule_wp_status": wp_st,
                    "wp_rest_base": schedule_rest_base,
                    "wp_schedule_error": "",
                    "wp_schedule_state": "queued",
                    "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "wp_generation_attempts": 0,
                    "wp_schedule_batch_id": batch_id,
                    "wp_schedule_batch_index": sched_idx,
                    "wp_schedule_batch_total": len(targets) - skipped,
                    "wp_schedule_tz_offset_min": "",
                },
            )
        )
        sched_idx += 1

    if bulk_updates:
        _storage.bulk_update_articles(bulk_updates)
    return len(bulk_updates), skipped, None


def _status_display_article(st: str | None) -> str:
    s = (st or "pending").strip().lower()
    if s == "published":
        return "Published"
    if s == "draft":
        return "Draft"
    return "Pending"


def _filter_articles_by_date_range(
    articles: list[dict], date_from: str | None, date_to: str | None
) -> list[dict]:
    """Filter by calendar date using posted_at when set, otherwise created_at."""
    df = (date_from or "").strip()
    dt = (date_to or "").strip()
    if not df and not dt:
        return articles
    d_from = None
    d_to = None
    if df:
        try:
            d_from = datetime.strptime(df, "%Y-%m-%d").date()
        except ValueError:
            pass
    if dt:
        try:
            d_to = datetime.strptime(dt, "%Y-%m-%d").date()
        except ValueError:
            pass
    out: list[dict] = []
    for a in articles:
        ref = (a.get("posted_at") or "").strip() or (a.get("created_at") or "").strip()
        adt = _parse_article_datetime(ref)
        if not adt:
            continue
        ad = adt.date()
        if d_from and ad < d_from:
            continue
        if d_to and ad > d_to:
            continue
        out.append(a)
    return out


def _article_status_key(a: dict) -> str:
    st = (a.get("status") or "pending").strip().lower()
    if st not in {"pending", "draft", "published"}:
        return "pending"
    return st


def _filter_articles_by_status(articles: list[dict], status_key: str | None) -> list[dict]:
    """Keep articles by status key (pending / draft / published / scheduled). Empty key = no filter."""
    sk = (status_key or "").strip().lower()
    if sk == "scheduled":
        out: list[dict] = []
        for a in articles:
            if (a.get("wp_post_id") or "").strip():
                continue
            if _article_wp_scheduled_at_str(a):
                out.append(a)
        return out
    if sk not in {"pending", "draft", "published"}:
        return articles
    return [a for a in articles if _article_status_key(a) == sk]


def _filter_articles_by_query(articles: list[dict], q: str | None) -> list[dict]:
    s = (q or "").strip().lower()
    if not s:
        return articles
    out: list[dict] = []
    for a in articles:
        title = (a.get("title") or "").strip().lower()
        fk = (a.get("focus_keyphrase") or "").strip().lower()
        kws = " ".join([str(x).strip().lower() for x in (a.get("keywords") or []) if str(x).strip()])
        if s in title or s in fk or s in kws:
            out.append(a)
    return out


def _build_articles_excel_bytes(articles: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Articles"
    headers = ["Article title", "Focus Keyphrase", "Targeting Keywords", "Live URL", "Status"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    for row_idx, a in enumerate(articles, start=2):
        row = [
            a.get("title") or "",
            (a.get("focus_keyphrase") or "").strip(),
            ", ".join(a.get("keywords") or []),
            (a.get("wp_link") or "").strip(),
            _status_display_article(a.get("status")),
        ]
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 44
    ws.column_dimensions["D"].width = 48
    ws.column_dimensions["E"].width = 14
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _normalize_bulk_sheet_header_cell(x) -> str:
    """Lowercase header labels and normalize spaces to underscores (e.g. 'Focus Keyphrase' -> 'focus_keyphrase')."""
    s = str(x).strip().lower() if x is not None else ""
    s = re.sub(r"[\s\-]+", "_", s)
    return s


def _build_bulk_upload_sample_bytes() -> bytes:
    """
    Sample sheet for bulk importing articles into a project.
    Columns: Title, Focus Keyphrase, Targeting Keywords (imported rows are always pending).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Sample"

    headers = ["Title", "Focus Keyphrase", "Targeting Keywords"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    examples = [
        [
            "Key Factors That Define Top Supreme Court Lawyers for Complex Litigation",
            "top supreme court lawyers india",
            "top supreme court lawyers india, senior advocate Supreme Court",
        ],
        [
            "Writ Petition under Article 226: Procedure and Reliefs",
            "writ petition article 226",
            "writ petition article 226, high court writ jurisdiction",
        ],
    ]
    for r, row in enumerate(examples, start=2):
        for cidx, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=cidx, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 45
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _normalize_website_url(raw: str) -> str:
    s = (raw or "").strip()
    if not s:
        raise ValueError("Website URL is required.")
    if not re.match(r"^https?://", s, re.IGNORECASE):
        s = "https://" + s
    if len(s) > 2048:
        raise ValueError("Website URL is too long.")
    return s.rstrip("/")


def _is_project_wordpress_configured(project: dict) -> bool:
    return bool((project.get("wp_username") or "").strip() and (project.get("wp_app_password") or "").strip())


def _wp_post_types_for_project(project: dict) -> tuple[list[dict[str, str]], str | None]:
    """Fetch REST post types from WordPress (GET /wp/v2/types), or a safe fallback."""
    fallback = [{"slug": "post", "name": "Posts", "rest_base": "posts"}]
    if not _is_project_wordpress_configured(project):
        return fallback, None
    try:
        cfg = WordPressConfig(
            site_url=(project.get("wp_site_url") or project.get("website_url") or "").strip(),
            username=(project.get("wp_username") or "").strip(),
            application_password=(project.get("wp_app_password") or "").strip(),
        )
        types = fetch_rest_post_types(cfg)
        if not types:
            return fallback, "WordPress returned no post types; using Posts only."
        return types, None
    except Exception as e:
        return fallback, str(e)


def _normalize_wp_rest_base(rest_base: str | None, allowed: set[str]) -> str:
    rb = (rest_base or "").strip().strip("/") or "posts"
    if rb in allowed:
        return rb
    if "posts" in allowed:
        return "posts"
    return next(iter(sorted(allowed))) if allowed else "posts"


def _parse_keywords(raw: str) -> list[str]:
    items = [x.strip() for x in (raw or "").split(",")]
    items = [x for x in items if x]
    # de-dup preserve order
    seen = set()
    out: list[str] = []
    for k in items:
        key = k.lower()
        if key in seen:
            continue
        seen.add(key)
        out.append(k)
    return out


def _dedup_case_insensitive(items: list[str]) -> list[str]:
    """De-dup strings preserving order (case-insensitive)."""
    seen: set[str] = set()
    out: list[str] = []
    for it in items or []:
        s = (it or "").strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return out


def _normalize_wp_tags(value: Any) -> list[str]:
    """
    Normalize a tags/keywords field into a list of tag names.
    Accepts list[str] or comma-separated string.
    """
    if value is None:
        return []
    if isinstance(value, list):
        return _dedup_case_insensitive([str(x).strip() for x in value if str(x).strip()])[:10]
    return _parse_keywords(str(value))[:10]


def _parse_focus_keyphrase_single_field(raw: str) -> tuple[str | None, str | None]:
    """
    Focus keyphrase must be one phrase (no comma-separated list).
    Returns (value or None if empty, error_message or None).
    """
    parts = _parse_keywords(raw or "")
    if len(parts) > 1:
        return None, "Focus keyphrase must be a single phrase (no commas)."
    if not parts:
        return None, None
    return parts[0], None


def _article_keywords_list(article: dict) -> list[str]:
    """Normalize stored article keywords to a list (max 10)."""
    kw = article.get("keywords") or []
    if isinstance(kw, list):
        out = _dedup_case_insensitive([str(x).strip() for x in kw if str(x).strip()])
    else:
        out = _parse_keywords(str(kw))
    return out[:10]


def _sanitize_filename(name: str) -> str:
    name = (name or "").strip() or "article"
    name = re.sub(r"[^a-zA-Z0-9\-_ ]+", "", name).strip()
    name = re.sub(r"\s+", "_", name)
    return (name[:80] or "article") + ".txt"


def _extract_first_json_object(text: str) -> dict:
    """
    Best-effort extraction of the first JSON object from an LLM response.
    """
    s = (text or "").strip()
    if not s:
        raise ValueError("Empty model response.")
    # Direct JSON
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass

    # Try to find a JSON object within text.
    start = s.find("{")
    end = s.rfind("}")
    if start != -1 and end != -1 and end > start:
        chunk = s[start : end + 1]
        obj = json.loads(chunk)
        if isinstance(obj, dict):
            return obj
    raise ValueError("Could not parse JSON from model response.")


def _get_openai_client_and_model(api_key: str | None):
    # Lazy import so the app can start even if deps missing; error shown on generate.
    from openai import OpenAI

    key = (api_key or "").strip()
    if not key:
        # Prefer explicit OpenAI key, then Groq key.
        key = os.environ.get("OPENAI_API_KEY", "").strip() or os.environ.get("GROQ_API_KEY", "").strip()
    if not key:
        raise ValueError(
            "Missing API key. Set OPENAI_API_KEY or GROQ_API_KEY, or paste a key in the form."
        )

    # If it's a Groq key (usually starts with gsk_) route to Groq's OpenAI-compatible endpoint.
    if key.lower().startswith("gsk_") or os.environ.get("LLM_PROVIDER", "").strip().lower() == "groq":
        base_url = os.environ.get("GROQ_BASE_URL", "https://api.groq.com/openai/v1").strip()
        model = os.environ.get("GROQ_MODEL", "llama-3.3-70b-versatile").strip() or "llama-3.3-70b-versatile"
        return OpenAI(api_key=key, base_url=base_url), model

    model = os.environ.get("OPENAI_MODEL", "gpt-4.1-mini").strip() or "gpt-4.1-mini"
    return OpenAI(api_key=key), model


def _can_run_openai_image_prompt_optimizer(api_key: str | None) -> bool:
    """ChatGPT prompt optimizer requires a real OpenAI key (not Groq-only)."""
    k = (api_key or "").strip() or os.environ.get("OPENAI_API_KEY", "").strip()
    if not k:
        return False
    if k.lower().startswith("gsk_"):
        return False
    if os.environ.get("LLM_PROVIDER", "").strip().lower() == "groq":
        return False
    return True


def _optimize_image_prompt_for_featured(
    raw_prompt: str,
    *,
    api_key: str | None,
    style: str,
    enabled: bool,
) -> tuple[str, str, str | None]:
    """
    Rewrite the interpolated image prompt for consistent semi-realistic (or other) style.
    Returns (final_prompt, optimizer_model_used, error_or_none).
    Skips or falls back to raw_prompt when disabled, unavailable, or on failure.
    """
    raw = (raw_prompt or "").strip()
    if not raw:
        return "", "", None
    if not enabled:
        return raw, "", None
    if not _can_run_openai_image_prompt_optimizer(api_key):
        return raw, "", None
    try:
        client, base_model = _get_openai_client_and_model(api_key)
        base_url = str(getattr(client, "base_url", "") or "")
        if "groq.com" in base_url.lower():
            return raw, "", None
    except Exception:
        return raw, "", None

    opt_model = (os.environ.get("OPENAI_IMAGE_PROMPT_MODEL") or "").strip() or base_model
    style_key = (style or "semi_real").strip().lower()
    style_desc = _IMAGE_STYLE_INSTRUCTIONS.get(style_key, _IMAGE_STYLE_INSTRUCTIONS["semi_real"])

    system = (
        "You rewrite image generation prompts for OpenAI's image generation API. "
        "Output a single English prompt string only—no markdown, no quotes, no preamble. "
        "Keep the user's subject and intent. Add concrete camera, composition, and lighting when helpful. "
        "Unless the user explicitly asks for text in the image, specify: no text, no watermarks, no logos."
    )
    user = f"""Style goal: {style_desc}

User prompt (placeholders already filled):
{raw}

Rewrite into one optimized prompt (max ~400 words)."""

    try:
        def _call():
            return client.chat.completions.create(
                model=opt_model,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user.strip()},
                ],
                temperature=0.45,
                max_tokens=700,
                timeout=_llm_timeout_seconds(),
            )

        resp = _with_retries(_call, tries=3, label="image prompt optimizer")
        out = (resp.choices[0].message.content or "").strip()
        if not out:
            raise ValueError("Empty optimizer response.")
        return out[:4000], opt_model, None
    except Exception as e:
        return raw, "", str(e)


def _generate_article_markdown(
    title: str,
    keywords: list[str],
    api_key: str | None,
    *,
    article_prompt: str | None = None,
    focus_keyphrase: str | None = None,
) -> str:
    client, model = _get_openai_client_and_model(api_key)

    kw = ", ".join(keywords) if keywords else "(none)"
    fk = (focus_keyphrase or "").strip()
    custom = (article_prompt or "").strip()
    if custom:
        system = (
            "You are an expert writer. Follow the user's instructions precisely. "
            "Produce well-structured Markdown suitable for web publishing unless instructed otherwise."
        )
        raw_template = custom
        interpolated = _interpolate_article_prompt_template(raw_template, title, keywords, focus_keyphrase)
        uses_placeholders = _prompt_template_has_placeholders(raw_template)
        if uses_placeholders:
            user = f"""{interpolated}

Do not use a Markdown # line (H1) in the article body; the site shows the post title as the only H1. Start sections with ## (H2) or deeper. You may begin with paragraphs or an H2 section—never repeat the title as a # heading."""
        else:
            fk_line = f"Focus keyphrase (Yoast): {fk}\n\n" if fk else ""
            user = f"""{interpolated}

{fk_line}Article/post title: {title}
Targeting keywords (use naturally where appropriate): {kw}

Do not use `#` (H1) in the Markdown. Start the first section with ## (H2) or deeper; the theme displays the title separately."""
    else:
        system = (
            "Role & Perspective: You are a senior legal research analyst, subject-matter expert in Indian law, "
            "and enterprise SEO/AEO/GEO strategist. You draft high-authority informational legal content for a law "
            "firm knowledge repository, strictly aligned with Bar Council of India regulations and non-solicitation norms."
        )
        user = f"""
Write a ~2,000-word, high-authority, informational legal article for a law firm's website knowledge repository.

Title: {title}
Targeting keywords (use naturally): {kw}

Requirements:
- Approx length: ~2,000 words (±10%).
- Output in Markdown (clean conversion to HTML).
- Do not use `#` (H1) in the Markdown body—the WordPress theme outputs the article title as the page H1. Start all section headings at `##` (H2) or deeper (###, ####).
- Use scannable formatting: short paragraphs, bullet points, numbered lists, and a table where appropriate.
- SEO/AEO/GEO: answer high-intent queries succinctly; include clear definitions, step-by-step legal explanations, and FAQ-style sections.
- Keyword integration:
  - Use the targeting keywords naturally (no stuffing).
  - Include short-tail, long-tail, and conversational queries.
  - Place keywords/synonyms in headings (H2–H4 only; never H1 in the body), intro, conclusion, and contextual body sections.
  - Include 2–3 additional external keywords relevant to the topic.
  - Include internal/external keywords aligned to the website context: https://sheokandlegal.com/
- Legal accuracy & Indian law framework:
  - Reference only accurate and current Indian laws (Constitution Articles, Central Acts/Codes/Rules/Regulations).
  - If a law is amended/replaced, reflect the updated framework.
  - Cite authoritative Supreme Court/High Court judgments where relevant (avoid speculative citations).
  - Keep statutory/judicial citations precise and contextual.
- Tone, style & compliance:
  - Formal, neutral, and authoritative.
  - Avoid promotional language, calls to action, outcome assurances, or legal advice.
  - Position as general legal information only.

Mandatory structure (use `##` headings in this order; do not use `#`):
1) Introduction – Context, scope, and relevance
2) Conceptual Overview – Foundational explanation
3) Statutory Framework Under Indian Law – Laws, sections, constitutional provisions
4) Rights, Duties, and Legal Obligations
5) Procedural Aspects and Legal Mechanisms
6) Judicial Interpretation and Landmark Case Laws
7) Practical Implications for Individuals and Businesses
8) Common Misconceptions and Clarifications
9) Frequently Asked Questions (AEO-Optimized) – 5–8 Q&As
10) Emerging Trends and Legal Developments in India
11) Conclusion – Key takeaways and informational summary
"""

    def _call():
        return client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user.strip()},
            ],
            temperature=0.7,
            timeout=_llm_timeout_seconds(),
        )

    resp = _with_retries(_call, tries=3, label="article generation")
    return (resp.choices[0].message.content or "").strip()


def _generate_yoast_fields(
    title: str,
    keywords: list[str],
    api_key: str | None,
    *,
    website_context_url: str | None = None,
    focus_keyphrase_preset: str | None = None,
) -> dict:
    client, model = _get_openai_client_and_model(api_key)

    kw = ", ".join(keywords) if keywords else "(none)"
    site = (website_context_url or "").strip() or "https://sheokandlegal.com/"
    preset = (focus_keyphrase_preset or "").strip()
    system = (
        "You are an enterprise SEO/AEO strategist for an Indian law firm website. "
        "Return STRICT JSON only (no markdown fences, no extra text)."
    )
    preset_block = ""
    if preset:
        preset_block = f"""
CRITICAL: The Yoast focus keyphrase is FIXED by the user. Use this exact string (copy verbatim, do not paraphrase):
"{preset}"
Build meta_title and meta_description around this keyphrase.
"""
    fk_rule = (
        '- focus_keyphrase: MUST be the exact fixed string given in CRITICAL above.\n'
        if preset
        else "- focus_keyphrase: 2-6 words, must match search intent, prefer one of the targeting keywords if suitable.\n"
    )
    user = f"""
Generate Yoast SEO fields for this article.

Title: {title}
Targeting keywords: {kw}
{preset_block}
Constraints:
{fk_rule}- meta_title: 50-60 characters target, include focus keyphrase near the start, readable.
- meta_description: 140-160 characters target, include focus keyphrase once, include a benefit + CTA, no quotes.
- Avoid promotional/salesy language; keep informational and compliant with Indian legal ethics.
- Website context for phrasing: {site}

Return strict JSON with exactly these keys:
{{"focus_keyphrase":"...","meta_title":"...","meta_description":"..."}}
"""

    def _call():
        return client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": system},
                {"role": "user", "content": user.strip()},
            ],
            temperature=0.4,
            timeout=_llm_timeout_seconds(),
        )

    resp = _with_retries(_call, tries=3, label="yoast fields")
    raw = (resp.choices[0].message.content or "").strip()
    obj = _extract_first_json_object(raw)
    out = {
        "focus_keyphrase": (obj.get("focus_keyphrase") or "").strip(),
        "meta_title": (obj.get("meta_title") or "").strip(),
        "meta_description": (obj.get("meta_description") or "").strip(),
    }
    if preset:
        out["focus_keyphrase"] = preset
    return out


@app.get("/")
def home():
    if not _is_authenticated():
        return render_template(
            "auth.html",
            storage_mode=_storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo",
            storage_init_error=_storage.storage_init_error() if hasattr(_storage, "storage_init_error") else None,
            account_deleted=(request.args.get("account_deleted") or "").strip(),
        )

    _maybe_backfill_orphan_project_owners()
    projects = _load_projects()
    current_user = _current_user() or {}
    is_admin = (current_user.get("role") or "").strip().lower() == "admin"
    section = (request.args.get("section") or "projects").strip().lower()
    if section not in {"projects", "users", "limits", "profile"}:
        section = "projects"
    if not is_admin:
        if section in {"users", "limits"}:
            section = "projects"
    # Client id/secret come from .env (loaded from the app directory at startup).
    cid = (os.environ.get("GOOGLE_OAUTH_CLIENT_ID") or "").strip()
    csec = (os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET") or "").strip()
    gsc_oauth_env_configured = bool(cid and csec)
    gsc_google_connected = False
    gsc_google_email = None
    gsc_google_libs_missing = False
    try:
        import google_integration as gi

        gsc_google_email = gi.get_stored_email()
        gsc_google_connected = gi.get_valid_credentials() is not None
    except ImportError:
        gsc_google_libs_missing = True

    users = []
    plans = {}
    if is_admin:
        try:
            users = _storage.list_users() if hasattr(_storage, "list_users") else []
        except Exception:
            users = []
        try:
            plans = _storage.load_plans() if hasattr(_storage, "load_plans") else {}
        except Exception:
            plans = {}
        try:
            users = [_admin_user_profile_snapshot(u) for u in users if isinstance(u, dict)]
        except Exception:
            # Fallback to basic user list if aggregation fails (e.g., DB not connected).
            pass
    _cu_home = _current_user()
    watch_tour = (request.args.get("watch_tour") or "").strip()
    show_onboarding_tour = bool((_cu_home and _cu_home.get("pending_product_tour")) or (watch_tour == "1"))
    return render_template(
        "home.html",
        projects=projects,
        max_projects=_MAX_PROJECTS,
        gsc_google_connected=gsc_google_connected,
        gsc_google_email=gsc_google_email,
        gsc_oauth_env_configured=gsc_oauth_env_configured,
        gsc_google_libs_missing=gsc_google_libs_missing,
        storage_mode=_storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo",
        storage_init_error=_storage.storage_init_error() if hasattr(_storage, "storage_init_error") else None,
        current_user=current_user,
        is_admin=is_admin,
        all_timezones=_all_timezones(),
        admin_section=section,
        home_section=section,
        admin_users=users,
        admin_plans=plans,
        show_onboarding_tour=show_onboarding_tour,
        primary_admin_email=_PRIMARY_PROTECTED_ADMIN_EMAIL,
        live_connect_meta=_LIVE_CONNECT_META,
    )


@app.get("/privacy")
def privacy_policy():
    # Public page: show signed-in context if available.
    current_user = _current_user() if _is_authenticated() else None
    is_admin = bool(current_user and (current_user.get("role") or "").strip().lower() == "admin")
    return render_template(
        "privacy.html",
        current_user=current_user,
        is_admin=is_admin,
    )


@app.get("/terms")
def terms_and_conditions():
    # Public page: show signed-in context if available.
    current_user = _current_user() if _is_authenticated() else None
    is_admin = bool(current_user and (current_user.get("role") or "").strip().lower() == "admin")
    return render_template(
        "terms.html",
        current_user=current_user,
        is_admin=is_admin,
    )


@app.post("/admin/users/subscription")
def admin_update_user_subscription():
    if not _is_authenticated() or (session.get("role") or "").strip().lower() != "admin":
        return redirect(url_for("home"))
    uid = (request.form.get("user_id") or "").strip()
    plan = (request.form.get("subscription_type") or "").strip().lower()
    if not uid or not plan:
        flash("Missing user or plan.", "error")
        return redirect(url_for("home", section="users"))
    try:
        plans = _storage.load_plans() if hasattr(_storage, "load_plans") else {}
        if plans and plan not in plans:
            flash("Unknown subscription plan.", "error")
            return redirect(url_for("home", section="users"))
        ok = _storage.update_user_fields(uid, {"subscription_type": plan}) if hasattr(_storage, "update_user_fields") else False
        if not ok:
            flash("User not found.", "error")
            return redirect(url_for("home", section="users"))
    except Exception as e:
        flash(f"Could not update user plan: {e}", "error")
        return redirect(url_for("home", section="users"))
    flash("Subscription updated.", "success")
    return redirect(url_for("home", section="users"))


@app.post("/admin/users/role")
def admin_update_user_role():
    if not _is_authenticated() or (session.get("role") or "").strip().lower() != "admin":
        return redirect(url_for("home"))
    uid = (request.form.get("user_id") or "").strip()
    role = (request.form.get("role") or "").strip().lower()
    if not uid or not role:
        flash("Missing user or role.", "error")
        return redirect(url_for("home", section="users"))
    if role not in {"user", "admin", "editor"}:
        flash("Invalid role.", "error")
        return redirect(url_for("home", section="users"))
    try:
        target = _storage.get_user_by_id(uid)
        if target and (target.get("email") or "").strip().lower() == _PRIMARY_PROTECTED_ADMIN_EMAIL:
            flash("The primary admin user role cannot be changed.", "error")
            return redirect(url_for("home", section="users"))
        ok = _storage.update_user_fields(uid, {"role": role}) if hasattr(_storage, "update_user_fields") else False
        if not ok:
            flash("User not found.", "error")
            return redirect(url_for("home", section="users"))
    except Exception as e:
        flash(f"Could not update user role: {e}", "error")
        return redirect(url_for("home", section="users"))
    flash("Role updated.", "success")
    return redirect(url_for("home", section="users"))


@app.post("/admin/plans/update")
def admin_update_plan():
    if not _is_authenticated() or (session.get("role") or "").strip().lower() != "admin":
        return redirect(url_for("home"))
    key = (request.form.get("plan_key") or "").strip().lower()
    if not key:
        flash("Missing plan.", "error")
        return redirect(url_for("home", section="limits"))
    if not re.fullmatch(r"[a-z0-9_]{2,40}", key or ""):
        flash("Plan key must be 2-40 chars: lowercase letters, numbers, underscore.", "error")
        return redirect(url_for("home", section="limits"))
    def _int(name: str, default: int) -> int:
        try:
            return int((request.form.get(name) or "").strip() or default)
        except Exception:
            return default
    payload = {
        "name": (request.form.get("plan_name") or key).strip()[:100],
        "max_projects": _int("max_projects", 2),
        "max_articles_per_day": _int("max_articles_per_day", 0),
        "max_articles_per_month": _int("max_articles_per_month", 0),
        "max_writing_prompts": _int("max_writing_prompts", 1),
        "writing_prompt_char_limit": _int("writing_prompt_char_limit", 4000),
        "max_image_prompts": _int("max_image_prompts", 1),
        "image_prompt_char_limit": _int("image_prompt_char_limit", 2000),
        "allow_scheduling": (request.form.get("allow_scheduling") or "") == "on",
        "max_scheduled_articles_per_month": _int("max_scheduled_articles_per_month", 0),
        "allow_export": (request.form.get("allow_export") or "") == "on",
        "allow_bulk_upload": (request.form.get("allow_bulk_upload") or "") == "on",
    }
    try:
        _storage.upsert_plan(key, payload) if hasattr(_storage, "upsert_plan") else None
    except Exception as e:
        flash(f"Could not save plan: {e}", "error")
        return redirect(url_for("home", section="limits"))
    flash("Plan saved.", "success")
    return redirect(url_for("home", section="limits"))


@app.post("/admin/projects/connect-live")
def admin_connect_live_projects():
    """Assign Sheokand Legal, KCS & TTSFM (by website URL host) to a user account."""
    if not _is_authenticated() or (session.get("role") or "").strip().lower() != "admin":
        flash("Unauthorized.", "error")
        return redirect(url_for("home"))
    if (request.form.get("confirm") or "").strip() != "1":
        flash("Please confirm that you want to connect these projects to the selected account.", "error")
        return redirect(url_for("home", section="limits"))
    email = (request.form.get("user_email") or "").strip().lower() or _PRIMARY_PROTECTED_ADMIN_EMAIL
    assigned, warnings = _assign_live_projects_to_user_email(email)
    fatal = False
    for w in warnings:
        low = w.lower()
        if any(
            x in low
            for x in (
                "no user found",
                "missing email",
                "not available",
                "could not load",
                "user id is missing",
            )
        ):
            flash(w, "error")
            fatal = True
        else:
            flash(w, "warning")
    if assigned:
        flash(f"Connected {len(assigned)} project(s) to {email}.", "success")
    elif not fatal and not warnings:
        flash("No projects matched the live site URLs (check website URLs on your projects).", "error")
    return redirect(url_for("home", section="limits"))


@app.post("/profile/update")
def update_profile():
    if not _is_authenticated():
        return redirect(url_for("home"))
    uid = (session.get("user_id") or "").strip()
    full_name = (request.form.get("full_name") or "").strip()
    phone = (request.form.get("phone") or "").strip()
    timezone = (request.form.get("timezone") or "").strip()
    try:
        upd = {"full_name": full_name, "phone": phone}
        if timezone:
            upd["timezone"] = timezone
        _storage.update_user_fields(uid, upd) if hasattr(_storage, "update_user_fields") else None
    except Exception as e:
        flash(f"Could not update profile: {e}", "error")
        return redirect(url_for("home", section="profile"))
    flash("Profile updated.", "success")
    return redirect(url_for("home", section="profile"))


@app.post("/account/delete")
def account_delete():
    if not _is_authenticated():
        return redirect(url_for("home"))
    pwd = request.form.get("password") or ""
    uid = (session.get("user_id") or "").strip()
    user = _storage.get_user_by_id(uid) if uid else None
    if not user or not (user.get("password_hash") or ""):
        flash("Could not verify account.", "error")
        return redirect(url_for("home", section="profile"))
    if not check_password_hash(user["password_hash"], pwd):
        flash("Incorrect password. Your account was not deleted.", "error")
        return redirect(url_for("home", section="profile"))
    if (user.get("email") or "").strip().lower() == _PRIMARY_PROTECTED_ADMIN_EMAIL:
        flash("This account cannot be deleted.", "error")
        return redirect(url_for("home", section="profile"))
    try:
        _purge_user_projects_and_files(uid)
        if not _storage.delete_user(uid):
            flash("Could not remove your account from the database.", "error")
            return redirect(url_for("home", section="profile"))
    except Exception as e:
        app.logger.exception("account delete")
        flash(f"Could not delete account: {e}", "error")
        return redirect(url_for("home", section="profile"))
    session.clear()
    return redirect(url_for("home", account_deleted="1"))


@app.post("/admin/users/delete")
def admin_delete_user():
    if not _is_authenticated() or (session.get("role") or "").strip().lower() != "admin":
        flash("Unauthorized.", "error")
        return redirect(url_for("home"))
    target_id = (request.form.get("user_id") or "").strip()
    if not target_id:
        flash("Missing user.", "error")
        return redirect(url_for("home", section="users"))
    target = _storage.get_user_by_id(target_id)
    if not target:
        flash("User not found.", "error")
        return redirect(url_for("home", section="users"))
    if (target.get("email") or "").strip().lower() == _PRIMARY_PROTECTED_ADMIN_EMAIL:
        flash("The primary admin account cannot be deleted.", "error")
        return redirect(url_for("home", section="users"))
    try:
        _purge_user_projects_and_files(target_id)
        if not _storage.delete_user(target_id):
            flash("Could not remove the user record.", "error")
            return redirect(url_for("home", section="users"))
    except Exception as e:
        app.logger.exception("admin delete user")
        flash(f"Could not delete user: {e}", "error")
        return redirect(url_for("home", section="users"))
    flash(f"User {(target.get('email') or target_id).strip()} and all related data were removed.", "success")
    curr = (session.get("user_id") or "").strip()
    if curr == target_id:
        session.clear()
        return redirect(url_for("home", account_deleted="1"))
    return redirect(url_for("home", section="users"))


@app.post("/auth/login")
def auth_login():
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    if not email or not password:
        flash("Email and password are required.", "error")
        return redirect(url_for("home"))
    user = _storage.get_user_by_email(email)
    if not user or not (user.get("password_hash") or ""):
        flash("Invalid email or password.", "error")
        return redirect(url_for("home"))
    if not check_password_hash(user["password_hash"], password):
        flash("Invalid email or password.", "error")
        return redirect(url_for("home"))
    session["user_id"] = user["id"]
    session["role"] = (user.get("role") or "user").strip().lower()
    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(user["id"], {"last_activity_at": _utc_now_str()})
    except Exception:
        pass
    flash("Logged in.", "success")
    fresh = _storage.get_user_by_id(user["id"])
    if fresh and fresh.get("pending_product_tour"):
        return redirect(url_for("home", section="projects"))
    return redirect(url_for("home"))


@app.post("/auth/register")
def auth_register():
    email = (request.form.get("email") or "").strip().lower()
    password = request.form.get("password") or ""
    if not email or not password:
        flash("Email and password are required.", "error")
        return redirect(url_for("home"))
    if len(password) < 8:
        flash("Password must be at least 8 characters.", "error")
        return redirect(url_for("home"))
    existing = _storage.get_user_by_email(email)
    if existing:
        flash("An account with that email already exists.", "error")
        return redirect(url_for("home"))
    new_uid = str(uuid.uuid4())
    try:
        _storage.insert_user(
            {
                "id": new_uid,
                "email": email,
                "password_hash": generate_password_hash(password),
                "role": "user",
                "subscription_type": "beta",
                "full_name": "",
                "phone": "",
                "last_activity_at": "",
                "pending_product_tour": True,
                "created_at": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
            }
        )
    except ValueError as ve:
        if "email" in str(ve).lower():
            flash("An account with that email already exists.", "error")
        else:
            flash("Could not create account.", "error")
        return redirect(url_for("home"))
    except Exception:
        flash("Could not create account (database unavailable).", "error")
        return redirect(url_for("home"))
    session["user_id"] = new_uid
    session["role"] = "user"
    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(new_uid, {"last_activity_at": _utc_now_str()})
    except Exception:
        pass
    flash("Welcome! Here is a quick tour of Auto Articles.", "success")
    return redirect(url_for("home", section="projects"))


@app.post("/auth/logout")
def auth_logout():
    session.clear()
    flash("Logged out.", "success")
    return redirect(url_for("home"))


@app.post("/account/onboarding-tour/complete")
def onboarding_tour_complete():
    if not _is_authenticated():
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    uid = (session.get("user_id") or "").strip()
    if not uid:
        return jsonify({"ok": False, "error": "unauthorized"}), 401
    try:
        if hasattr(_storage, "update_user_fields"):
            _storage.update_user_fields(uid, {"pending_product_tour": False})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500
    return jsonify({"ok": True})


@app.post("/projects")
def add_project():
    if not _is_authenticated():
        return redirect(url_for("home"))
    name = (request.form.get("project_name") or "").strip()
    url_raw = (request.form.get("website_url") or "").strip()
    if not name:
        flash("Please enter a project name.", "error")
        return redirect(url_for("home"))
    try:
        url = _normalize_website_url(url_raw)
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("home"))
    projects = _load_projects()
    limits = _plan_limit_summary(_current_user())
    plan_max = limits.get("max_projects", _MAX_PROJECTS)
    try:
        eff_max = min(int(plan_max), int(_MAX_PROJECTS))
    except Exception:
        eff_max = _MAX_PROJECTS
    if len(projects) >= eff_max:
        if eff_max != _MAX_PROJECTS:
            flash(
                f"Your plan allows a maximum of {eff_max} project(s). Upgrade or remove a project to add another.",
                "error",
            )
        else:
            flash(f"Maximum {_MAX_PROJECTS} projects reached (3×3 grid). Remove a project to add another.", "error")
        return redirect(url_for("home"))
    new_id = str(uuid.uuid4())
    _storage.insert_project(
        {
            "id": new_id,
            "owner_user_id": (session.get("user_id") or "").strip(),
            "name": name[:200],
            "website_url": url,
            "wp_site_url": url,
            "wp_username": "",
            "wp_app_password": "",
            "wp_category_ids": "",
            "prompts": [],
            "default_prompt_id": "",
            "image_prompts": [],
            "default_image_prompt_id": "",
            "image_style": "semi_real",
            "optimize_image_prompt": True,
            "context_links": [],
            "gsc_property_url": "",
            "gsc_index_on_publish": True,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    )
    flash(f"Project “{name}” added. Configure WordPress in Project settings.", "success")
    return redirect(url_for("project_detail", project_id=new_id, open_settings=1))


@app.post("/projects/<project_id>/settings")
def update_project_settings(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))

    wp_site_url_raw = (request.form.get("wp_site_url") or "").strip() or project.get("website_url") or ""
    wp_username = (request.form.get("wp_username") or "").strip()
    wp_app_password = (request.form.get("wp_app_password") or "").strip()
    wp_category_ids = (request.form.get("wp_category_ids") or "").strip()

    try:
        wp_site_url = _normalize_website_url(wp_site_url_raw)
    except ValueError as e:
        flash(str(e), "error")
        return redirect(url_for("project_detail", project_id=project_id))

    # Category IDs are optional but if provided must be comma-separated integers
    if wp_category_ids:
        for part in wp_category_ids.split(","):
            part = part.strip()
            if not part:
                continue
            if not part.isdigit():
                flash("Category IDs must be comma-separated numbers (e.g. 12, 34).", "error")
                return redirect(url_for("project_detail", project_id=project_id))

    ok = _update_project_fields(
        project_id,
        {
            "wp_site_url": wp_site_url,
            "wp_username": wp_username,
            "wp_app_password": wp_app_password,
            "wp_category_ids": wp_category_ids,
        },
    )
    if not ok:
        flash("Could not save project settings to MongoDB. Check the server log and your connection.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    flash("Project settings saved.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/featured-image-settings")
def update_project_featured_image_settings(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    style = (request.form.get("image_style") or "semi_real").strip().lower()
    if style not in ("semi_real", "photorealistic", "illustration"):
        style = "semi_real"
    optimize = (request.form.get("optimize_image_prompt") or "") == "on"
    _update_project_fields(
        project_id,
        {
            "image_style": style,
            "optimize_image_prompt": optimize,
        },
    )
    flash("Featured image settings saved.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/prompts/add")
def add_project_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    name = (request.form.get("prompt_name") or "").strip()
    text = (request.form.get("prompt_text") or "").strip()
    if not name:
        flash("Please enter a prompt name.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not text:
        flash("Please enter the prompt text used for article generation.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_prompts(proj)
    prompts = list(proj.get("prompts") or [])
    limits = _plan_limit_summary(_current_user())
    if len(prompts) >= int(limits.get("max_writing_prompts", 9999)):
        flash(
            f"Your plan allows a maximum of {limits.get('max_writing_prompts')} writing prompt(s) per project.",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    if len(text) > int(limits.get("writing_prompt_char_limit", 100_000)):
        flash(
            f"Writing prompt is too long (max {limits.get('writing_prompt_char_limit')} characters).",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    new_id = str(uuid.uuid4())
    prompts.append({"id": new_id, "name": name[:200], "text": text[:100_000]})
    updates: dict = {"prompts": prompts}
    set_default = (request.form.get("set_as_default") or "") == "on"
    if len(prompts) == 1 or set_default:
        updates["default_prompt_id"] = new_id
    _update_project_fields(project_id, updates)
    flash(f"Prompt “{name[:80]}” added.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/prompts/default")
def set_project_default_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("prompt_id") or "").strip()
    proj = dict(project)
    _normalize_project_prompts(proj)
    if not prompt_id or not _get_prompt_by_id(proj, prompt_id):
        flash("Select a valid prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    _update_project_fields(project_id, {"default_prompt_id": prompt_id})
    flash("Default prompt updated for this project.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/prompts/update")
def update_project_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("prompt_id") or "").strip()
    name = (request.form.get("prompt_name") or "").strip()
    text = (request.form.get("prompt_text") or "").strip()
    if not prompt_id:
        flash("Missing prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not name:
        flash("Please enter a prompt name.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not text:
        flash("Please enter the prompt text.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_prompts(proj)
    prompts = list(proj.get("prompts") or [])
    limits = _plan_limit_summary(_current_user())
    if len(text) > int(limits.get("writing_prompt_char_limit", 100_000)):
        flash(
            f"Writing prompt is too long (max {limits.get('writing_prompt_char_limit')} characters).",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    found = False
    for i, p in enumerate(prompts):
        if isinstance(p, dict) and (p.get("id") or "") == prompt_id:
            prompts[i] = {"id": prompt_id, "name": name[:200], "text": text[:100_000]}
            found = True
            break
    if not found:
        flash("Prompt not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    updates: dict = {"prompts": prompts}
    if (request.form.get("set_as_default") or "") == "on":
        updates["default_prompt_id"] = prompt_id
    _update_project_fields(project_id, updates)
    flash(f"Prompt “{name[:80]}” updated.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/prompts/delete")
def delete_project_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("prompt_id") or "").strip()
    if not prompt_id:
        flash("Missing prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_prompts(proj)
    prompts = [p for p in (proj.get("prompts") or []) if isinstance(p, dict) and (p.get("id") or "") != prompt_id]
    if len(prompts) == len(proj.get("prompts") or []):
        flash("Prompt not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    updates: dict = {"prompts": prompts}
    default_id = (proj.get("default_prompt_id") or "").strip()
    if default_id == prompt_id:
        updates["default_prompt_id"] = (prompts[0].get("id") or "") if prompts else ""
    _update_project_fields(project_id, updates)
    flash("Prompt removed.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/image-prompts/add")
def add_project_image_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    name = (request.form.get("image_prompt_name") or "").strip()
    text = (request.form.get("image_prompt_text") or "").strip()
    if not name:
        flash("Please enter an image prompt name.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not text:
        flash("Please enter the image prompt text.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_image_prompts(proj)
    prompts = list(proj.get("image_prompts") or [])
    limits = _plan_limit_summary(_current_user())
    if len(prompts) >= int(limits.get("max_image_prompts", 9999)):
        flash(
            f"Your plan allows a maximum of {limits.get('max_image_prompts')} image prompt(s) per project.",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    if len(text) > int(limits.get("image_prompt_char_limit", 100_000)):
        flash(
            f"Image prompt is too long (max {limits.get('image_prompt_char_limit')} characters).",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    new_id = str(uuid.uuid4())
    prompts.append({"id": new_id, "name": name[:200], "text": text[:100_000]})
    updates: dict = {"image_prompts": prompts}
    set_default = (request.form.get("set_as_default") or "") == "on"
    if len(prompts) == 1 or set_default:
        updates["default_image_prompt_id"] = new_id
    _update_project_fields(project_id, updates)
    flash(f"Image prompt “{name[:80]}” added.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/image-prompts/default")
def set_project_default_image_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("image_prompt_id") or "").strip()
    proj = dict(project)
    _normalize_project_image_prompts(proj)
    if not prompt_id or not _get_image_prompt_by_id(proj, prompt_id):
        flash("Select a valid image prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    _update_project_fields(project_id, {"default_image_prompt_id": prompt_id})
    flash("Default image prompt updated.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/wp-defaults")
def set_project_wp_defaults(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))

    proj = dict(project)
    wp_types, _ = _wp_post_types_for_project(proj)
    allowed = {t["rest_base"] for t in wp_types}
    rest_base = _normalize_wp_rest_base(request.form.get("default_wp_rest_base"), allowed)

    st = (request.form.get("default_wp_status") or "draft").strip().lower()
    if st not in ("draft", "publish"):
        st = "draft"

    _update_project_fields(project_id, {"default_wp_rest_base": rest_base, "default_wp_status": st})
    flash("Default WordPress post type and status saved.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/image-prompts/update")
def update_project_image_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("image_prompt_id") or "").strip()
    name = (request.form.get("image_prompt_name") or "").strip()
    text = (request.form.get("image_prompt_text") or "").strip()
    if not prompt_id:
        flash("Missing prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not name:
        flash("Please enter a prompt name.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not text:
        flash("Please enter the prompt text.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_image_prompts(proj)
    prompts = list(proj.get("image_prompts") or [])
    limits = _plan_limit_summary(_current_user())
    if len(text) > int(limits.get("image_prompt_char_limit", 100_000)):
        flash(
            f"Image prompt is too long (max {limits.get('image_prompt_char_limit')} characters).",
            "error",
        )
        return redirect(url_for("project_detail", project_id=project_id, open_settings=1))
    found = False
    for i, p in enumerate(prompts):
        if isinstance(p, dict) and (p.get("id") or "") == prompt_id:
            prompts[i] = {"id": prompt_id, "name": name[:200], "text": text[:100_000]}
            found = True
            break
    if not found:
        flash("Prompt not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    updates: dict = {"image_prompts": prompts}
    if (request.form.get("set_as_default") or "") == "on":
        updates["default_image_prompt_id"] = prompt_id
    _update_project_fields(project_id, updates)
    flash(f"Image prompt “{name[:80]}” updated.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/image-prompts/delete")
def delete_project_image_prompt(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prompt_id = (request.form.get("image_prompt_id") or "").strip()
    if not prompt_id:
        flash("Missing prompt.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_image_prompts(proj)
    prompts = [p for p in (proj.get("image_prompts") or []) if isinstance(p, dict) and (p.get("id") or "") != prompt_id]
    if len(prompts) == len(proj.get("image_prompts") or []):
        flash("Prompt not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    updates: dict = {"image_prompts": prompts}
    default_id = (proj.get("default_image_prompt_id") or "").strip()
    if default_id == prompt_id:
        updates["default_image_prompt_id"] = (prompts[0].get("id") or "") if prompts else ""
    _update_project_fields(project_id, updates)
    flash("Image prompt removed.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


def _validate_context_link_phrase(phrase: str) -> str | None:
    p = (phrase or "").strip()
    if not p:
        return "Text to match is required."
    if len(p) > 2000:
        return "Text is too long (max 2000 characters)."
    if "]" in p or "\n" in p or "\r" in p:
        return "Text cannot contain ] or line breaks."
    return None


@app.post("/projects/<project_id>/context-links/add")
def add_project_context_link(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    phrase = (request.form.get("link_phrase") or request.form.get("context_link_phrase") or "").strip()
    url = (request.form.get("link_url") or request.form.get("context_link_url") or "").strip()
    err = _validate_context_link_phrase(phrase)
    if err:
        flash(err, "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not _is_valid_context_link_url(url):
        flash("Enter a valid http(s) URL (no ) in the URL).", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_context_links(proj)
    links = list(proj.get("context_links") or [])
    new_id = str(uuid.uuid4())
    links.append({"id": new_id, "phrase": phrase[:2000], "url": url[:2048]})
    _update_project_fields(project_id, {"context_links": links})
    flash("Context link added.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/context-links/update")
def update_project_context_link(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    link_id = (request.form.get("context_link_id") or "").strip()
    phrase = (request.form.get("link_phrase") or "").strip()
    url = (request.form.get("link_url") or "").strip()
    if not link_id:
        flash("Missing link id.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    err = _validate_context_link_phrase(phrase)
    if err:
        flash(err, "error")
        return redirect(url_for("project_detail", project_id=project_id))
    if not _is_valid_context_link_url(url):
        flash("Enter a valid http(s) URL (no ) in the URL).", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_context_links(proj)
    links = list(proj.get("context_links") or [])
    found = False
    for i, item in enumerate(links):
        if isinstance(item, dict) and (item.get("id") or "") == link_id:
            links[i] = {"id": link_id, "phrase": phrase[:2000], "url": url[:2048]}
            found = True
            break
    if not found:
        flash("Link not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    _update_project_fields(project_id, {"context_links": links})
    flash("Context link updated.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/context-links/delete")
def delete_project_context_link(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    link_id = (request.form.get("context_link_id") or "").strip()
    if not link_id:
        flash("Missing link id.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    proj = dict(project)
    _normalize_project_context_links(proj)
    links = [x for x in (proj.get("context_links") or []) if isinstance(x, dict) and (x.get("id") or "") != link_id]
    if len(links) == len(proj.get("context_links") or []):
        flash("Link not found.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    _update_project_fields(project_id, {"context_links": links})
    flash("Context link removed.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.get("/oauth/google/start")
def google_oauth_start():
    try:
        import google_integration as gi
    except ImportError:
        flash(
            "Install Google client libraries: pip install google-api-python-client google-auth-oauthlib google-auth-httplib2",
            "error",
        )
        return redirect(url_for("home"))
    if not gi.oauth_client_configured():
        flash("Set GOOGLE_OAUTH_CLIENT_ID and GOOGLE_OAUTH_CLIENT_SECRET in the environment.", "error")
        return redirect(url_for("home"))
    redirect_uri = url_for("google_oauth_callback", _external=True)
    flow = gi.build_flow(redirect_uri)
    state = secrets.token_hex(16)
    session["oauth_google_state"] = state
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        prompt="consent",
        state=state,
        include_granted_scopes="true",
    )
    return redirect(authorization_url)


@app.get("/oauth/google/callback")
def google_oauth_callback():
    try:
        import google_integration as gi
    except ImportError:
        flash("Google libraries are not installed.", "error")
        return redirect(url_for("home"))
    err = request.args.get("error")
    if err:
        flash(request.args.get("error_description") or "Google sign-in was cancelled or denied.", "error")
        return redirect(url_for("home"))
    if request.args.get("state") != session.pop("oauth_google_state", None):
        flash("Invalid OAuth state. Try connecting again.", "error")
        return redirect(url_for("home"))
    redirect_uri = url_for("google_oauth_callback", _external=True)
    flow = gi.build_flow(redirect_uri)
    try:
        flow.fetch_token(authorization_response=request.url)
    except Exception as e:
        flash(f"Could not complete Google sign-in: {e}", "error")
        return redirect(url_for("home"))
    creds = flow.credentials
    email = gi.fetch_user_email(creds)
    gi.save_oauth_session(creds, email)
    flash(
        f"Google account connected{f' ({email})' if email else ''}. Assign a Search Console property on each project as needed.",
        "success",
    )
    return redirect(url_for("home"))


@app.post("/oauth/google/disconnect")
def google_oauth_disconnect():
    try:
        import google_integration as gi

        gi.disconnect()
        flash("Google account disconnected from this app.", "success")
    except ImportError:
        pass
    return redirect(request.referrer or url_for("home"))


@app.post("/projects/<project_id>/gsc")
def update_project_gsc(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    prop = (request.form.get("gsc_property_url") or "").strip()
    index_on = (request.form.get("gsc_index_on_publish") or "") == "on"
    try:
        import google_integration as gi

        creds = gi.get_valid_credentials()
        if prop and creds:
            sites = {
                s.get("siteUrl")
                for s in gi.list_search_console_sites(creds)
                if isinstance(s, dict)
            }
            if prop not in sites:
                flash("That Search Console property is not in your connected account.", "error")
                return redirect(url_for("project_detail", project_id=project_id))
        elif prop and not creds:
            flash("Connect your Google account from the home page before choosing a property.", "error")
            return redirect(url_for("project_detail", project_id=project_id))
    except ImportError:
        pass
    _update_project_fields(
        project_id,
        {
            "gsc_property_url": prop,
            "gsc_index_on_publish": index_on,
        },
    )
    flash("Search Console settings saved for this project.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.post("/projects/<project_id>/delete")
def delete_project(project_id: str):
    pid = (project_id or "").strip()
    if not pid:
        flash("Invalid project.", "error")
        return redirect(url_for("home"))
    # Enforce ownership / role scope before deleting.
    project = _get_project_by_id(pid)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    if not _storage.delete_project_and_articles(pid):
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    flash("Project removed.", "success")
    return redirect(url_for("home"))


@app.get("/projects/<project_id>")
def project_detail(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    bulk_err = _pop_bulk_schedule_error(project_id)
    if bulk_err:
        flash(bulk_err, "error")
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    status_filter = request.args.get("status", "").strip().lower()
    if status_filter not in {"pending", "draft", "published", "scheduled"}:
        status_filter = ""
    q = (request.args.get("q") or "").strip()
    tab = (request.args.get("tab") or "articles").strip().lower()
    if tab not in {"articles", "configuration", "tools", "scheduled"}:
        tab = "articles"
    # Avoid loading all articles across the DB; fetch only this project's listing fields.
    try:
        all_articles = (
            _storage.load_articles_listing_for_project(project_id, limit=5000)
            if hasattr(_storage, "load_articles_listing_for_project")
            else _articles_for_project(project_id)
        )
    except Exception:
        all_articles = _articles_for_project(project_id)
    article_count_total = len(all_articles)
    all_articles.sort(key=lambda a: (a.get("created_at") or ""), reverse=True)
    filtered_all = _filter_articles_by_date_range(all_articles, date_from, date_to)
    filtered_all = _filter_articles_by_status(filtered_all, status_filter)
    filtered_all = _filter_articles_by_query(filtered_all, q)

    per_page = 10
    try:
        page = int((request.args.get("page") or "").strip() or "1")
    except Exception:
        page = 1
    if page < 1:
        page = 1
    total_filtered = len(filtered_all)
    total_pages = max(1, (total_filtered + per_page - 1) // per_page)
    if page > total_pages:
        page = total_pages
    start_i = (page - 1) * per_page
    end_i = start_i + per_page
    articles = filtered_all[start_i:end_i]
    proj = dict(project)
    _normalize_project_prompts(proj)
    _normalize_project_image_prompts(proj)
    _normalize_project_wp_defaults(proj)
    _normalize_project_context_links(proj)
    _normalize_project_gsc(proj)
    wp_post_types, wp_post_types_error = _wp_post_types_for_project(proj)
    gsc_sites: list[dict] = []
    gsc_google_connected = False
    gsc_oauth_env_configured = bool(
        (os.environ.get("GOOGLE_OAUTH_CLIENT_ID") or "").strip()
        and (os.environ.get("GOOGLE_OAUTH_CLIENT_SECRET") or "").strip()
    )
    gsc_google_email = None
    gsc_google_libs_missing = False
    try:
        import google_integration as gi

        gsc_google_email = gi.get_stored_email()
        creds = gi.get_valid_credentials()
        gsc_google_connected = creds is not None
        if creds:
            try:
                gsc_sites = gi.list_search_console_sites(creds)
            except Exception as e:
                app.logger.warning("Could not list Google Search Console sites: %s", e)
                gsc_sites = []
    except ImportError:
        gsc_google_libs_missing = True
    bulk_schedule_meta: list[dict] = []
    for a in articles:
        aid = (a.get("id") or "").strip()
        if not aid:
            continue
        bulk_schedule_meta.append(
            {
                "id": aid,
                "created_at": (a.get("created_at") or "").strip(),
                "hasBody": bool(a.get("hasBody")) if "hasBody" in a else bool((a.get("article") or "").strip()),
                "hasImage": bool(os.path.isfile(_article_featured_image_path(aid))),
            }
        )
    open_settings = request.args.get("open_settings", "").strip().lower() in ("1", "true", "yes")
    watch_sched = request.args.get("watch_schedule", "").strip().lower() in ("1", "true", "yes")
    poll_articles_wp = any(
        _article_wp_scheduled_at_str(a)
        for a in articles
    ) or any(
        (a.get("wp_post_id") and (a.get("status") or "").strip().lower() == "pending")
        for a in articles
    ) or any(
        ((a.get("posted_at") or "").strip() and (a.get("status") or "").strip().lower() == "pending")
        for a in articles
    )
    poll_schedule_ui = bool(poll_articles_wp or watch_sched)
    scheduled_pending_articles = _pending_scheduled_article_rows(project_id)
    try:
        counts = (
            _storage.count_articles_by_project_ids([project_id])
            if hasattr(_storage, "count_articles_by_project_ids")
            else None
        )
    except Exception:
        counts = None
    if counts:
        _pending_n = int(counts.get("pending") or 0)
        _draft_n = int(counts.get("draft") or 0)
        _published_n = int(counts.get("published") or 0)
    else:
        _pending_n = 0
        _draft_n = 0
        _published_n = 0
        for a in all_articles:
            st = (a.get("status") or "pending").strip().lower()
            if st == "published":
                _published_n += 1
            elif st == "draft":
                _draft_n += 1
            else:
                _pending_n += 1
    article_stats = {
        "total": article_count_total,
        "pending": _pending_n,
        "draft": _draft_n,
        "published": _published_n,
        "scheduled_queue": len(scheduled_pending_articles),
    }

    def _status_filter_qs(st: str) -> str:
        st = st.strip().lower()
        if st not in {"pending", "draft", "published"}:
            st = ""
        return urlencode(
            {k: v for k, v in [("date_from", date_from), ("date_to", date_to), ("status", st)] if v}
        )

    article_status_filter_qs = {
        "all": _status_filter_qs(""),
        "pending": _status_filter_qs("pending"),
        "draft": _status_filter_qs("draft"),
        "published": _status_filter_qs("published"),
    }
    return render_template(
        "project.html",
        project=proj,
        articles=articles,
        pagination={
            "page": page,
            "per_page": per_page,
            "total_items": total_filtered,
            "total_pages": total_pages,
            "start_index": 0 if total_filtered == 0 else (start_i + 1),
            "end_index": min(end_i, total_filtered),
        },
        date_from=date_from,
        date_to=date_to,
        q=q,
        active_tab=tab,
        article_count_total=article_count_total,
        bulk_schedule_meta=bulk_schedule_meta,
        scheduled_pending_articles=scheduled_pending_articles,
        status_filter=status_filter,
        article_status_filter_qs=article_status_filter_qs,
        project_settings_incomplete=not _project_wp_credentials_configured(proj),
        open_project_settings_modal=open_settings,
        poll_articles_wp=poll_articles_wp,
        poll_schedule_ui=poll_schedule_ui,
        wp_post_types=wp_post_types,
        wp_post_types_error=wp_post_types_error,
        gsc_sites=gsc_sites,
        gsc_google_connected=gsc_google_connected,
        gsc_oauth_env_configured=gsc_oauth_env_configured,
        gsc_google_email=gsc_google_email,
        gsc_google_libs_missing=gsc_google_libs_missing,
        article_stats=article_stats,
        plan_limits=_plan_limit_summary(_current_user()),
        plan_ui=_plan_ui_for_project(proj, _current_user()),
        storage_mode=_storage.storage_mode() if hasattr(_storage, "storage_mode") else "mongo",
        storage_init_error=_storage.storage_init_error() if hasattr(_storage, "storage_init_error") else None,
    )


@app.get("/projects/<project_id>/articles/status-summary")
def project_articles_status_summary(project_id: str):
    """JSON for refreshing article status/posted/scheduled columns without a full page reload."""
    project = _get_project_by_id(project_id)
    if not project:
        return jsonify({"error": "not found"}), 404
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    status_filter = request.args.get("status", "").strip().lower()
    if status_filter not in {"pending", "draft", "published"}:
        status_filter = ""
    # Fast path: avoid loading all articles from the DB.
    try:
        all_articles = (
            _storage.load_articles_for_project_minimal(
                project_id,
                status=status_filter or None,
                date_from=date_from or None,
                date_to=date_to or None,
                limit=600,
            )
            if hasattr(_storage, "load_articles_for_project_minimal")
            else _articles_for_project(project_id)
        )
    except Exception:
        all_articles = _articles_for_project(project_id)
    # Keep existing filtering logic (string date comparisons are tricky in Mongo).
    all_articles.sort(key=lambda a: (a.get("created_at") or ""), reverse=True)
    articles = _filter_articles_by_date_range(all_articles, date_from, date_to)
    if not hasattr(_storage, "load_articles_for_project_minimal") or not status_filter:
        articles = _filter_articles_by_status(articles, status_filter)
    out: list[dict] = []
    for a in articles:
        aid = (a.get("id") or "").strip()
        if not aid:
            continue
        st = (a.get("status") or "pending").strip().lower()
        if st not in {"pending", "draft", "published"}:
            st = "pending"
        gs = (a.get("gsc_status") or "pending").strip().lower()
        if gs == "requested":
            gs = "inspected"
        if gs not in {"pending", "inspected"}:
            gs = "pending"
        out.append(
            {
                "id": aid,
                "status": st,
                "posted_at": (a.get("posted_at") or "").strip(),
                "wp_scheduled_at": _article_wp_scheduled_at_str(a),
                "wp_schedule_error": (a.get("wp_schedule_error") or "").strip(),
                "gsc_status": gs,
                "gsc_inspection_requested_at": (a.get("gsc_inspection_requested_at") or "").strip(),
                "gsc_inspection_error": (a.get("gsc_inspection_error") or "").strip(),
            }
        )
    scheduled_pending = _pending_scheduled_article_rows(project_id)
    pipeline_error = _pop_bulk_schedule_error(project_id)
    return jsonify(
        {
            "articles": out,
            "scheduled_pending": scheduled_pending,
            "scheduled_queue": len(scheduled_pending),
            "pipeline_error": pipeline_error,
        }
    )


@app.get("/projects/<project_id>/articles/export")
def export_project_articles(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    limits = _plan_limit_summary(_current_user())
    if not limits.get("allow_export", True):
        flash("Upgrade plan to enable Export articles.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    date_from = request.args.get("date_from", "").strip()
    date_to = request.args.get("date_to", "").strip()
    status_filter = request.args.get("status", "").strip().lower()
    if status_filter not in {"pending", "draft", "published"}:
        status_filter = ""
    all_articles = _articles_for_project(project_id)
    all_articles.sort(key=lambda a: (a.get("created_at") or ""), reverse=True)
    filtered = _filter_articles_by_date_range(all_articles, date_from, date_to)
    filtered = _filter_articles_by_status(filtered, status_filter)
    safe_name = re.sub(r"[^a-zA-Z0-9\-_]+", "_", project.get("name") or "project")[:60] or "project"
    filename = f"{safe_name}_articles.xlsx"
    data = _build_articles_excel_bytes(filtered)
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/projects/<project_id>/articles/bulk/sample")
def bulk_upload_sample(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    limits = _plan_limit_summary(_current_user())
    if not limits.get("allow_bulk_upload", True):
        flash("Upgrade plan to enable Bulk Upload.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    safe_name = re.sub(r"[^a-zA-Z0-9\-_]+", "_", project.get("name") or "project")[:60] or "project"
    filename = f"{safe_name}_bulk_upload_sample.xlsx"
    data = _build_bulk_upload_sample_bytes()
    return Response(
        data,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/projects/<project_id>/wordpress/plugin.zip")
def download_wordpress_plugin(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))

    try:
        import sys
        backend_root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "backend")
        if backend_root not in sys.path:
            sys.path.insert(0, backend_root)
        from app.services.wordpress_plugin_packager import build_plugin_zip_bytes, get_plugin_version
    except ImportError as e:
        return Response(f"Plugin packager unavailable: {e}", status=500, mimetype="text/plain")

    try:
        data, filename = build_plugin_zip_bytes()
        version = get_plugin_version()
    except Exception as e:
        return Response(f"Plugin bundle error: {e}", status=500, mimetype="text/plain")

    return Response(
        data,
        mimetype="application/zip",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store, no-cache, must-revalidate",
            "Pragma": "no-cache",
            "X-Riviso-Plugin-Version": version,
            "X-Riviso-Plugin-Slug": "riviso-content-operations",
        },
    )


@app.post("/projects/<project_id>/articles/bulk/upload")
def bulk_upload_articles(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    limits = _plan_limit_summary(_current_user())
    if not limits.get("allow_bulk_upload", True):
        flash("Upgrade plan to enable Bulk Upload.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    file = request.files.get("file")
    if not file or not getattr(file, "filename", ""):
        flash("Please choose an Excel file to upload.", "error")
        return redirect(url_for("project_detail", project_id=project_id))

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValueError("The uploaded sheet is empty.")

        header = [_normalize_bulk_sheet_header_cell(x) for x in rows[0]]

        def col_idx(*names: str) -> int | None:
            for name in names:
                if name in header:
                    return header.index(name)
            return None

        idx_title = col_idx("title")
        idx_focus = col_idx("focus_keyphrase")
        idx_keywords = col_idx("targeting_keywords", "keywords")
        if idx_title is None:
            raise ValueError("Missing required column: title")

        to_add: list[dict] = []
        errors: list[str] = []

        for r_i, row in enumerate(rows[1:], start=2):
            title = ""
            if idx_title < len(row) and row[idx_title] is not None:
                title = str(row[idx_title]).strip()
            if not title:
                # skip completely blank rows
                if all((c is None or str(c).strip() == "") for c in row):
                    continue
                errors.append(f"Row {r_i}: title is required.")
                continue

            fk_raw = ""
            if idx_focus is not None and idx_focus < len(row) and row[idx_focus] is not None:
                fk_raw = str(row[idx_focus]).strip()
            fk_val, fk_err = _parse_focus_keyphrase_single_field(fk_raw)
            if fk_err:
                errors.append(f"Row {r_i}: {fk_err}")
                continue

            kw_raw = ""
            if idx_keywords is not None and idx_keywords < len(row) and row[idx_keywords] is not None:
                kw_raw = str(row[idx_keywords]).strip()
            keywords = _parse_keywords(kw_raw)
            if len(keywords) > 10:
                errors.append(f"Row {r_i}: maximum 10 keywords allowed.")
                continue

            to_add.append(
                {
                    "id": str(uuid.uuid4()),
                    "project_id": project_id,
                    "title": title[:500],
                    "keywords": keywords,
                    "status": "pending",
                    "article": "",
                    "focus_keyphrase": (fk_val or "")[:500],
                    "meta_title": "",
                    "meta_description": "",
                    "generated_at": "",
                    "posted_at": "",
                    "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "gsc_status": "pending",
                }
            )

        if errors:
            flash("Bulk upload failed: " + " ".join(errors[:6]) + (" ..." if len(errors) > 6 else ""), "error")
            return redirect(url_for("project_detail", project_id=project_id))

        if not to_add:
            flash("No valid rows found to import.", "error")
            return redirect(url_for("project_detail", project_id=project_id))

        _storage.insert_articles_batch(to_add)
        flash(f"Imported {len(to_add)} articles.", "success")
        return redirect(url_for("project_detail", project_id=project_id))
    except Exception as e:
        flash(f"Bulk upload failed: {e}", "error")
        return redirect(url_for("project_detail", project_id=project_id))


def _project_detail_redirect_query_from_parts(date_from: str, date_to: str, filter_status: str) -> str:
    df = (date_from or "").strip()
    dt = (date_to or "").strip()
    fs = (filter_status or "").strip().lower()
    if fs not in {"pending", "draft", "published"}:
        fs = ""
    return urlencode({k: v for k, v in [("date_from", df), ("date_to", dt), ("status", fs)] if v})


@app.post("/projects/<project_id>/articles/status-batch")
def batch_update_article_statuses(project_id: str):
    """Apply multiple article status changes in one request (JSON)."""
    project = _get_project_by_id(project_id)
    if not project:
        return jsonify({"ok": False, "error": "Project not found."}), 404
    data = request.get_json(silent=True) or {}
    updates = data.get("updates")
    if not isinstance(updates, list) or not updates:
        return jsonify({"ok": False, "error": "No updates."}), 400

    changed = 0
    for u in updates:
        if not isinstance(u, dict):
            continue
        aid = (u.get("id") or "").strip()
        raw = (u.get("status") or "").strip().lower()
        if raw not in {"pending", "draft", "published"}:
            continue
        article = _get_article_by_id(aid)
        if not article or (article.get("project_id") or "") != project_id:
            continue
        _update_article_fields(aid, {"status": raw})
        changed += 1

    if changed == 0:
        return jsonify({"ok": False, "error": "No valid articles updated."}), 400

    df = (data.get("date_from") or "").strip()
    dt = (data.get("date_to") or "").strip()
    fs = (data.get("filter_status") or "").strip().lower()
    q = _project_detail_redirect_query_from_parts(df, dt, fs)
    base = url_for("project_detail", project_id=project_id)
    redirect_url = base + ("?" + q if q else "")
    return jsonify({"ok": True, "changed": changed, "redirect": redirect_url})


@app.post("/projects/<project_id>/articles/<article_id>/status")
def update_article_status(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))
    raw = (request.form.get("status") or "").strip().lower()
    if raw not in {"pending", "draft", "published"}:
        flash("Invalid status.", "error")
        df = request.form.get("date_from", "").strip()
        dt = request.form.get("date_to", "").strip()
        fs = request.form.get("filter_status", "").strip().lower()
        q = _project_detail_redirect_query_from_parts(df, dt, fs)
        url = url_for("project_detail", project_id=project_id)
        return redirect(url + ("?" + q if q else ""))
    _update_article_fields(article_id, {"status": raw})
    flash("Status updated.", "success")
    df = request.form.get("date_from", "").strip()
    dt = request.form.get("date_to", "").strip()
    fs = request.form.get("filter_status", "").strip().lower()
    q = _project_detail_redirect_query_from_parts(df, dt, fs)
    url = url_for("project_detail", project_id=project_id)
    return redirect(url + ("?" + q if q else ""))


@app.post("/projects/<project_id>/articles/<article_id>/schedule/cancel")
def cancel_article_schedule(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))
    limits = _plan_limit_summary(_current_user())
    if not limits.get("allow_scheduling", True):
        flash("Upgrade plan to enable Scheduling.", "error")
        return _redirect_project_detail_with_dates(project_id)
    if article.get("wp_post_id"):
        flash("This article is already posted to WordPress.", "error")
        return _redirect_project_detail_with_dates(project_id)
    if not _article_wp_scheduled_at_str(article):
        flash("Nothing to cancel — this article is not scheduled.", "error")
        return _redirect_project_detail_with_dates(project_id)
    _update_article_fields(article_id, _cleared_wp_schedule_fields())
    flash("Scheduled post cancelled for this article.", "success")
    return _redirect_project_detail_with_dates(project_id)


@app.post("/projects/<project_id>/articles/<article_id>/schedule/update")
def update_article_schedule(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))
    limits = _plan_limit_summary(_current_user())
    if not limits.get("allow_scheduling", True):
        flash("Upgrade plan to enable Scheduling.", "error")
        return _redirect_project_detail_with_dates(project_id)
    if article.get("wp_post_id"):
        flash("This article is already posted to WordPress.", "error")
        return _redirect_project_detail_with_dates(project_id)
    if not _article_wp_scheduled_at_str(article):
        flash("This article is not scheduled.", "error")
        return _redirect_project_detail_with_dates(project_id)
    if not _is_project_wordpress_configured(project):
        flash("Configure WordPress for this project before editing a schedule.", "error")
        return _redirect_project_detail_with_dates(project_id)

    raw_dt = (request.form.get("schedule_at") or "").strip()
    dt = _parse_bulk_schedule_datetime(raw_dt)
    if not dt:
        flash("Please enter a valid date and time.", "error")
        return _redirect_project_detail_with_dates(project_id)
    owner = _storage.get_user_by_id((project.get("owner_user_id") or "").strip()) if hasattr(_storage, "get_user_by_id") else None
    tz_name = _user_timezone_name(owner)
    dt_local = dt.replace(microsecond=0)
    dt_utc_str = _local_to_utc_str(dt_local, tz_name)
    try:
        dt_utc = datetime.strptime(dt_utc_str, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        dt_utc = dt_local
    dt_utc = _clamp_future_schedule_dt(dt_utc)
    sched_str = dt_local.strftime("%Y-%m-%d %H:%M:%S")
    sched_utc_str = dt_utc.strftime("%Y-%m-%d %H:%M:%S")

    wp_st = (request.form.get("schedule_wp_status") or "draft").strip().lower()
    if wp_st not in ("draft", "publish"):
        wp_st = "draft"

    _update_article_fields(
        article_id,
        {
            "wp_scheduled_at": sched_str,
            "wp_scheduled_at_utc": sched_utc_str,
            "wp_schedule_wp_status": wp_st,
            "wp_schedule_error": "",
            "wp_schedule_batch_id": "",
            "wp_schedule_batch_index": "",
            "wp_schedule_batch_total": "",
            "wp_schedule_tz_offset_min": "",
        },
    )
    flash("Schedule updated.", "success")
    return _redirect_project_detail_with_dates(project_id)


@app.post("/projects/<project_id>/articles/bulk-action")
def bulk_articles_action(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    action = (request.form.get("action") or "").strip().lower()
    ids = [x.strip() for x in request.form.getlist("article_ids") if x.strip()]
    if not ids:
        flash("No articles selected.", "error")
        return _redirect_project_detail_with_dates(project_id)

    arts = _load_articles()
    id_set = set(ids)
    targets = [
        a
        for a in arts
        if (a.get("id") or "") in id_set and (a.get("project_id") or "") == project_id
    ]
    if not targets:
        flash("No matching articles for this project.", "error")
        return _redirect_project_detail_with_dates(project_id)

    if action == "delete":
        del_ids = [(t.get("id") or "").strip() for t in targets if (t.get("id") or "").strip()]
        for t in targets:
            _delete_article_featured_image_file((t.get("id") or "").strip())
        _storage.delete_articles_by_ids(del_ids)
        flash(f"Deleted {len(targets)} article(s).", "success")
        return _redirect_project_detail_with_dates(project_id)

    if action == "change_status":
        raw = (request.form.get("new_status") or "").strip().lower()
        if raw not in {"pending", "draft", "published"}:
            flash("Invalid status.", "error")
            return _redirect_project_detail_with_dates(project_id)
        for t in targets:
            _update_article_fields((t.get("id") or "").strip(), {"status": raw})
        flash(f"Updated status for {len(targets)} article(s).", "success")
        return _redirect_project_detail_with_dates(project_id)

    if action == "schedule":
        limits = _plan_limit_summary(_current_user())
        if not limits.get("allow_scheduling", True):
            flash("Upgrade plan to enable Scheduling.", "error")
            return _redirect_project_detail_with_dates(project_id)
        if not _is_project_wordpress_configured(project):
            flash("Configure WordPress for this project before scheduling posts.", "error")
            return _redirect_project_detail_with_dates(project_id)

        # Monthly scheduling cap (how many articles can be scheduled in the plan renewal period).
        sched_limit = int(limits.get("max_scheduled_articles_per_month") or 0)
        if sched_limit > 0:
            new_to_schedule = 0
            for t in targets:
                if (t.get("wp_post_id") or "").strip():
                    continue
                if _article_wp_scheduled_at_str(t):
                    continue
                new_to_schedule += 1
            if new_to_schedule > 0:
                ok_sched, sched_err = _schedule_quota_try_consume((session.get("user_id") or "").strip(), new_to_schedule)
                if not ok_sched:
                    flash(sched_err or "Monthly scheduling limit reached. Upgrade plan to continue.", "error")
                    return _redirect_project_detail_with_dates(project_id)
        wp_st = (request.form.get("schedule_wp_status") or "draft").strip().lower()
        if wp_st not in ("draft", "publish"):
            wp_st = "draft"

        proj_check = dict(project)
        _normalize_project_image_prompts(proj_check)
        need_img = len(proj_check.get("image_prompts") or []) > 0
        need_any_generation = False
        for t in targets:
            aid = (t.get("id") or "").strip()
            if not aid or t.get("wp_post_id"):
                continue
            need_body = not (t.get("article") or "").strip()
            need_img_file = need_img and not os.path.isfile(_article_featured_image_path(aid))
            if not need_body and not need_img_file:
                continue
            need_any_generation = True
            if not (t.get("title") or "").strip():
                flash(
                    "Cannot generate or schedule: one or more selected articles have no title. Add a title first.",
                    "error",
                )
                return _redirect_project_detail_with_dates(project_id)

        form_snapshot = {
            "schedule_wp_status": wp_st,
            "schedule_wp_rest_base": (request.form.get("schedule_wp_rest_base") or "").strip(),
            "bulk_prompt_id": (request.form.get("bulk_prompt_id") or "").strip(),
            "bulk_image_prompt_id": (request.form.get("bulk_image_prompt_id") or "").strip(),
            "schedule_times_json": request.form.get("schedule_times_json") or "",
            "client_tz_offset_min": (request.form.get("client_tz_offset_min") or "").strip(),
        }
        # Persist schedule metadata immediately so the Scheduled posts section updates right away.
        n_sched, skipped, sched_err = _bulk_set_wp_schedule_fields_now(project_id, list(ids), form_snapshot)
        if sched_err:
            flash(sched_err, "error")
            return _redirect_project_detail_with_dates(project_id)
        if not need_any_generation:
            msg = "Articles added to the WordPress schedule queue. Due posts are sent automatically when their time is reached."
            if skipped:
                msg += f" Skipped {skipped} (already posted)."
            flash(msg, "success")
            return _redirect_project_detail_with_dates(project_id, watch_schedule=True)

        threading.Thread(
            target=_bulk_schedule_thread_entry,
            args=(project_id, list(ids), form_snapshot),
            daemon=True,
        ).start()
        flash(
            "Generating content in the background. Scheduled items appear immediately; WordPress posting will run automatically when each time is due. "
            "If something fails, an error will appear when you open this project again. "
            "Past times are adjusted to a few minutes from now.",
            "success",
        )
        return _redirect_project_detail_with_dates(project_id, watch_schedule=True)

    flash("Unknown bulk action.", "error")
    return _redirect_project_detail_with_dates(project_id)


@app.post("/projects/<project_id>/articles")
def add_project_article(project_id: str):
    project = _get_project_by_id(project_id)
    if not project:
        flash("Project not found.", "error")
        return redirect(url_for("home"))
    title = (request.form.get("title") or "").strip()
    keywords_raw = request.form.get("keywords") or ""
    if not title:
        flash("Please enter an article title.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    keywords = _parse_keywords(keywords_raw)
    if len(keywords) > 10:
        flash("Maximum 10 targeting keywords allowed.", "error")
        return redirect(url_for("project_detail", project_id=project_id))
    fk_val, fk_err = _parse_focus_keyphrase_single_field(request.form.get("focus_keyphrase") or "")
    if fk_err:
        flash(fk_err, "error")
        return redirect(url_for("project_detail", project_id=project_id))
    _storage.insert_article(
        {
            "id": str(uuid.uuid4()),
            "project_id": project_id,
            "title": title[:500],
            "keywords": keywords,
            "status": "pending",
            "article": "",
            "focus_keyphrase": (fk_val or "")[:500],
            "meta_title": "",
            "meta_description": "",
            "generated_at": "",
            "posted_at": "",
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "gsc_status": "pending",
        }
    )
    flash("Article added.", "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.get("/projects/<project_id>/articles/<article_id>")
def article_edit(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))
    last = _article_to_last(article)
    proj = dict(project)
    _normalize_project_prompts(proj)
    _normalize_project_wp_defaults(proj)
    default_prompt = _get_prompt_by_id(proj, (proj.get("default_prompt_id") or "").strip())
    _normalize_project_image_prompts(proj)
    default_image_prompt = _get_image_prompt_by_id(proj, (proj.get("default_image_prompt_id") or "").strip())
    aid = (article.get("id") or "").strip()
    has_featured_image = bool(aid and os.path.isfile(_article_featured_image_path(aid)))
    requires_featured_image = len(proj.get("image_prompts") or []) > 0
    wp_post_types, wp_post_types_error = _wp_post_types_for_project(proj)
    allowed_bases = {t["rest_base"] for t in wp_post_types}
    project_default_wp_rest_base = _normalize_wp_rest_base((proj.get("default_wp_rest_base") or "posts"), allowed_bases)
    selected_wp_rest_base = _normalize_wp_rest_base((article.get("wp_rest_base") or project_default_wp_rest_base), allowed_bases)
    wp_status_default = (proj.get("default_wp_status") or "draft").strip().lower()
    if wp_status_default not in ("draft", "publish"):
        wp_status_default = "draft"
    selected_wp_status = (article.get("wp_last_wp_status") or "").strip().lower() or wp_status_default
    if selected_wp_status not in ("draft", "publish"):
        selected_wp_status = wp_status_default
    return render_template(
        "article_edit.html",
        project=proj,
        article=article,
        last=last,
        default_prompt_name=(default_prompt.get("name") if default_prompt else None),
        default_prompt_id=(proj.get("default_prompt_id") or "").strip(),
        default_image_prompt_name=(default_image_prompt.get("name") if default_image_prompt else None),
        default_image_prompt_id=(proj.get("default_image_prompt_id") or "").strip(),
        has_featured_image=has_featured_image,
        requires_featured_image=requires_featured_image,
        wp_post_types=wp_post_types,
        wp_post_types_error=wp_post_types_error,
        selected_wp_rest_base=selected_wp_rest_base,
        project_default_wp_rest_base=project_default_wp_rest_base,
        selected_wp_status=selected_wp_status,
        project_default_wp_status=wp_status_default,
        project_settings_incomplete=not _project_wp_credentials_configured(proj),
    )


@app.get("/projects/<project_id>/articles/<article_id>/featured-image.png")
def serve_article_featured_image(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        return Response("Not found", 404)
    path = _article_featured_image_path(article_id)
    if not os.path.isfile(path):
        return Response("Not found", 404)
    resp = send_file(path, mimetype="image/png")
    resp.headers["Cache-Control"] = "no-cache, max-age=0"
    return resp


def _generate_article_content_core(
    project: dict,
    article_id: str,
    *,
    title: str,
    keywords: list[str],
    writing_prompt_id: str,
    image_prompt_id: str,
    user_focus_keyphrase: str | None = None,
) -> tuple[bool, str | None, str | None]:
    """
    Shared article + optional featured image generation (same as article edit Generate).
    Returns (ok, error_message, image_error_message). image_error_message is set when
    article saved but OpenAI image generation failed (ok is still True).
    """
    api_key: str | None = None
    proj = dict(project)
    _normalize_project_prompts(proj)
    _normalize_project_image_prompts(proj)
    optimize_flag = bool(proj.get("optimize_image_prompt", True))
    image_style = (proj.get("image_style") or "semi_real").strip()
    form_prompt_id = (writing_prompt_id or "").strip()
    if form_prompt_id:
        sel = _get_prompt_by_id(proj, form_prompt_id)
        if not sel:
            return False, "Invalid writing prompt selection.", None
        article_prompt = (sel.get("text") or "").strip() or None
    else:
        article_prompt = _resolve_default_prompt_text(proj)
    site_url = (project.get("website_url") or project.get("wp_site_url") or "").strip() or None
    fk_preset = (user_focus_keyphrase or "").strip() or None

    image_prompts_list = proj.get("image_prompts") or []
    form_image_prompt_id = (image_prompt_id or "").strip()
    image_prompt_raw: str | None = None
    used_image_prompt_id = ""
    if image_prompts_list:
        if form_image_prompt_id:
            ip = _get_image_prompt_by_id(proj, form_image_prompt_id)
            if not ip:
                return False, "Invalid image prompt selection.", None
            image_prompt_raw = (ip.get("text") or "").strip() or None
            used_image_prompt_id = form_image_prompt_id
        else:
            image_prompt_raw = _resolve_default_image_prompt_text(proj)
            did = (proj.get("default_image_prompt_id") or "").strip()
            if did:
                used_image_prompt_id = did
        if not image_prompt_raw:
            return (
                False,
                "This project has image prompts but none could be used. "
                "Set a default image prompt on the project page or pick prompts in the schedule dialog.",
                None,
            )

    run_image = bool(image_prompts_list and image_prompt_raw)
    interpolated_image_prompt: str | None = None
    if run_image and image_prompt_raw:
        interpolated_image_prompt = _interpolate_article_prompt_template(
            image_prompt_raw, title, keywords, fk_preset
        )

    _delete_article_featured_image_file(article_id)

    article_md = ""
    yoast: dict = {}
    image_bytes: bytes | None = None
    image_err: str | None = None
    gen_meta_full: dict[str, Any] = {}

    try:
        # Keep peak memory low on small instances (e.g. Render free/starter):
        # generate article + yoast in parallel, but run image generation after.
        with ThreadPoolExecutor(max_workers=2) as ex:
            fut_article = ex.submit(
                _generate_article_markdown,
                title,
                keywords,
                api_key,
                article_prompt=article_prompt,
                focus_keyphrase=fk_preset,
            )
            fut_yoast = ex.submit(
                _generate_yoast_fields,
                title,
                keywords,
                api_key,
                website_context_url=site_url,
                focus_keyphrase_preset=fk_preset,
            )
            article_md = fut_article.result()
            yoast = fut_yoast.result()
    except Exception as e:
        return False, str(e), None

    if run_image and interpolated_image_prompt:
        final_prompt, opt_mod, opt_err = _optimize_image_prompt_for_featured(
            interpolated_image_prompt,
            api_key=api_key,
            style=image_style,
            enabled=optimize_flag,
        )
        try:
            image_bytes, gen_meta = _generate_featured_image_png_bytes(final_prompt)
            gen_meta_full = dict(gen_meta)
            gen_meta_full["featured_image_prompt_raw"] = interpolated_image_prompt
            gen_meta_full["featured_image_optimizer_model"] = (opt_mod or "") if optimize_flag else ""
            gen_meta_full["featured_image_prompt_optimizer_error"] = (opt_err or "") if optimize_flag else ""
        except Exception as e:
            image_err = str(e)

    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    fi_time = ""
    if run_image and image_bytes:
        try:
            _save_article_featured_image_png(article_id, image_bytes)
            fi_time = gen_time
        except Exception as e:
            image_err = image_err or str(e)
    elif run_image and image_err:
        pass

    fk_stored = (yoast.get("focus_keyphrase") or "").strip()
    if fk_preset:
        fk_stored = fk_preset
    upd: dict = {
        "title": title[:500],
        "keywords": keywords,
        "article": (article_md or "").strip(),
        "focus_keyphrase": fk_stored,
        "meta_title": (yoast.get("meta_title") or "").strip(),
        "meta_description": (yoast.get("meta_description") or "").strip(),
        "generated_at": gen_time,
        "status": "pending",
        "featured_image_generated_at": fi_time,
        "featured_image_prompt_id": used_image_prompt_id if (run_image and image_bytes) else "",
        "featured_image_source": ("generated" if (run_image and image_bytes) else ""),
    }
    if not run_image:
        upd["featured_image_generated_at"] = ""
        upd["featured_image_prompt_id"] = ""
        upd["featured_image_source"] = ""
        upd.update(_clear_featured_image_generation_meta_updates())
    elif run_image and image_bytes:
        upd.update(gen_meta_full)
    elif run_image and not image_bytes:
        upd.update(_clear_featured_image_generation_meta_updates())

    _update_article_fields(article_id, upd)
    return True, None, image_err


@app.post("/projects/<project_id>/articles/<article_id>/generate")
def generate_project_article(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))

    title = (request.form.get("title") or "").strip()
    keywords_raw = request.form.get("keywords") or ""
    keywords = _parse_keywords(keywords_raw)

    if not title:
        flash("Please enter an article title.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    if not keywords:
        flash("Please enter at least one targeting keyword (comma-separated, max 10).", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    if len(keywords) > 10:
        flash("Maximum 10 targeting keywords allowed.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    fk_val, fk_err = _parse_focus_keyphrase_single_field(request.form.get("focus_keyphrase") or "")
    if fk_err:
        flash(fk_err, "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))
    user_fk = (fk_val or "").strip() or None

    # Plan quota: number of articles that can be generated per day/month.
    ok_quota, quota_err = _article_quota_try_consume((session.get("user_id") or "").strip(), 1)
    if not ok_quota:
        flash((quota_err or "Current limit is exhausted. Upgrade plan to continue."), "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    proj = dict(project)
    ok, err, image_err = _generate_article_content_core(
        proj,
        article_id,
        title=title,
        keywords=keywords,
        writing_prompt_id=(request.form.get("prompt_id") or "").strip(),
        image_prompt_id=(request.form.get("image_prompt_id") or "").strip(),
        user_focus_keyphrase=user_fk,
    )
    if not ok:
        flash(f"Generation failed: {err}", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    run_image = len((proj.get("image_prompts") or [])) > 0
    if run_image and not image_err:
        flash("Article and featured image generated. You can post to WordPress.", "success")
    elif run_image and image_err:
        flash(f"Article generated, but featured image failed: {image_err}", "error")
    else:
        flash("Article generated. You can post to WordPress as Draft or Published.", "success")
    return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))


@app.post("/projects/<project_id>/articles/<article_id>/save-body")
def save_article_body(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        if request.headers.get("X-Requested-With") == "XMLHttpRequest":
            return jsonify({"ok": False, "error": "not_found"}), 404
        flash("Article not found.", "error")
        return redirect(url_for("home"))

    body = request.form.get("article_body")
    if body is None:
        body = ""
    ok_save = _update_article_fields(article_id, {"article": body[:500_000]})

    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"
    if wants_json:
        if not ok_save:
            return jsonify({"ok": False, "error": "save_failed"}), 500
        return jsonify({"ok": True})
    if not ok_save:
        flash("Could not save article to MongoDB.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))
    flash("Article saved.", "success")
    return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))


@app.post("/projects/<project_id>/articles/<article_id>/featured-image/clear")
def clear_article_featured_image(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))

    _delete_article_featured_image_file(article_id)
    _update_article_fields(
        article_id,
        {
            "featured_image_generated_at": "",
            "featured_image_prompt_id": "",
            "featured_image_source": "",
            **_clear_featured_image_generation_meta_updates(),
        },
    )
    flash("Featured image removed. You can generate a new one or upload one image.", "success")
    return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))


@app.post("/projects/<project_id>/articles/<article_id>/featured-image/upload")
def upload_article_featured_image(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))

    if os.path.isfile(_article_featured_image_path(article_id)):
        flash("Clear the current featured image before uploading a new one.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    f = request.files.get("file")
    if not f or not getattr(f, "filename", ""):
        flash("Choose one image file to upload.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    raw = f.read()
    if len(raw) > _MAX_FEATURED_UPLOAD_BYTES:
        flash(f"Image is too large (max { _MAX_FEATURED_UPLOAD_BYTES // (1024 * 1024) } MB).", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    try:
        png = _convert_upload_to_png_bytes(raw)
    except Exception as e:
        flash(f"Could not read that image. Use PNG, JPEG, WebP, or GIF. ({e})", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    try:
        _save_article_featured_image_png(article_id, png)
    except OSError as e:
        flash(f"Could not save image: {e}", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    _update_article_fields(
        article_id,
        {
            "featured_image_generated_at": ts,
            "featured_image_prompt_id": "",
            "featured_image_source": "uploaded",
            **_clear_featured_image_generation_meta_updates(),
        },
    )
    flash("Featured image saved. It will be used when you post to WordPress.", "success")
    return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))


def _execute_wordpress_post_from_last(last: dict) -> tuple[bool, str | None, dict]:
    """
    Uses request.form. Returns (ok, error_message, info) where info may include
    post_id, link, wp_status in {"draft", "publish"}.
    """
    if not last or not last.get("article"):
        return False, "Generate an article first before posting to WordPress.", {}

    site_url = (request.form.get("wp_site_url") or "").strip()
    username = (request.form.get("wp_username") or "").strip()
    app_password = (request.form.get("wp_app_password") or "").strip()
    remember = (request.form.get("wp_remember") or "") == "on"

    status = (request.form.get("wp_status") or "draft").strip().lower()
    if status not in {"draft", "publish"}:
        status = "draft"

    title = (request.form.get("wp_title") or "").strip() or (last.get("title") or "").strip()
    tags_raw = request.form.get("wp_tags") or ""
    tags = _parse_keywords(tags_raw)

    cat_raw = (request.form.get("wp_category_ids") or "").strip()
    category_ids: list[int] = []
    if cat_raw:
        for part in cat_raw.split(","):
            part = part.strip()
            if not part:
                continue
            if part.isdigit():
                category_ids.append(int(part))

    excerpt = (request.form.get("wp_excerpt") or "").strip() or None
    focus_keyphrase = (request.form.get("wp_focus_keyphrase") or "").strip() or (last.get("focus_keyphrase") or "").strip() or None
    meta_title = (request.form.get("wp_meta_title") or "").strip() or (last.get("meta_title") or "").strip() or None
    meta_description = (request.form.get("wp_meta_description") or "").strip() or (last.get("meta_description") or "").strip() or None

    if remember:
        session["wp_site_url"] = site_url
        session["wp_username"] = username
        session["wp_app_password"] = app_password
        session["wp_focus_keyphrase"] = focus_keyphrase or ""
        session["wp_meta_title"] = meta_title or ""
        session["wp_meta_description"] = meta_description or ""

    try:
        cfg = WordPressConfig(site_url=site_url, username=username, application_password=app_password)
        html = _article_body_to_wp_html(last["article"], None)
        if not html:
            raise ValueError("Generated article was empty after formatting.")

        tag_ids = ensure_tag_ids(cfg, tags) if tags else []
        yoast_meta = {}
        if focus_keyphrase:
            yoast_meta["_yoast_wpseo_focuskw"] = focus_keyphrase
        if meta_title:
            yoast_meta["_yoast_wpseo_title"] = meta_title
        if meta_description:
            yoast_meta["_yoast_wpseo_metadesc"] = meta_description

        created = create_post(
            cfg,
            title=title or "Untitled",
            html_content=html,
            status=status,
            excerpt=excerpt,
            tag_ids=tag_ids,
            category_ids=category_ids,
            meta=yoast_meta or None,
        )

        return True, None, {
            "post_id": created.get("id"),
            "link": created.get("link") or "",
            "wp_status": status,
        }
    except Exception as e:
        return False, str(e), {}


def _post_article_to_wordpress(
    project: dict,
    article: dict,
    article_id: str,
    *,
    wp_status: str = "draft",
    rest_base_preference: str | None = None,
    site_url_override: str | None = None,
    allow_without_featured_image: bool = False,
) -> tuple[bool, str | None, dict]:
    """
    Post one article to WordPress using project credentials.
    Returns (ok, error_message, info) with post_id, link, wp_status, rest_base on success.
    """
    last = _article_to_last(article)
    if not (last.get("article") or "").strip():
        return False, "Article body is empty.", {}

    if not _is_project_wordpress_configured(project):
        return False, "WordPress is not configured for this project.", {}

    proj_check = dict(project)
    _normalize_project_image_prompts(proj_check)
    need_featured = len(proj_check.get("image_prompts") or []) > 0
    has_file = os.path.isfile(_article_featured_image_path(article_id))
    if need_featured and not has_file and not allow_without_featured_image:
        return False, "Featured image is required before posting (generate, upload, or confirm posting without an image).", {}

    site_url = (site_url_override or "").strip() or (project.get("wp_site_url") or project.get("website_url") or "").strip()
    status = wp_status if wp_status in {"draft", "publish"} else "draft"
    tags = _normalize_wp_tags(last.get("keywords"))

    category_ids: list[int] = []
    cat_raw = (project.get("wp_category_ids") or "").strip()
    if cat_raw:
        for part in cat_raw.split(","):
            part = part.strip()
            if part.isdigit():
                category_ids.append(int(part))

    excerpt = None
    focus_keyphrase = (last.get("focus_keyphrase") or "").strip() or None
    meta_title = (last.get("meta_title") or "").strip() or None
    meta_description = (last.get("meta_description") or "").strip() or None

    wp_types, _ = _wp_post_types_for_project(project)
    allowed_bases = {t["rest_base"] for t in wp_types}
    rest_base = _normalize_wp_rest_base(rest_base_preference if rest_base_preference is not None else article.get("wp_rest_base"), allowed_bases)

    try:
        cfg = WordPressConfig(
            site_url=site_url,
            username=(project.get("wp_username") or "").strip(),
            application_password=(project.get("wp_app_password") or "").strip(),
        )
        html = _article_body_to_wp_html(last["article"], project)
        if not html:
            raise ValueError("Generated article was empty after formatting.")

        tag_ids = ensure_tag_ids(cfg, tags) if tags else []
        yoast_meta = {}
        if focus_keyphrase:
            yoast_meta["_yoast_wpseo_focuskw"] = focus_keyphrase
        if meta_title:
            yoast_meta["_yoast_wpseo_title"] = meta_title
        if meta_description:
            yoast_meta["_yoast_wpseo_metadesc"] = meta_description

        featured_media_id: int | None = None
        img_path = _article_featured_image_path(article_id)
        if os.path.isfile(img_path):
            with open(img_path, "rb") as img_f:
                img_bytes = img_f.read()
            safe_fn = re.sub(r"[^a-zA-Z0-9._-]+", "-", (last.get("title") or "featured")[:80]).strip("-") or "featured"
            media = upload_media(cfg, img_bytes, safe_fn + ".png")
            mid = media.get("id")
            if isinstance(mid, int):
                featured_media_id = mid
            else:
                # If the file exists but WP returns no numeric media ID, treat as failure when images are required.
                # Otherwise we'd create the post without a featured image and the user wouldn't know why.
                if need_featured and not allow_without_featured_image:
                    raise ValueError(f"WordPress media upload returned no numeric id: {mid!r}")

        created = create_post(
            cfg,
            title=(last.get("title") or "Untitled"),
            html_content=html,
            status=status,
            excerpt=excerpt,
            tag_ids=tag_ids,
            category_ids=category_ids,
            meta=yoast_meta or None,
            featured_media=featured_media_id,
            rest_base=rest_base,
        )
        wp_status_actual = (created.get("status") or status or "draft")
        if isinstance(wp_status_actual, str):
            wp_status_actual = wp_status_actual.strip().lower()
        else:
            wp_status_actual = str(wp_status_actual).lower()
        return True, None, {
            "post_id": created.get("id"),
            "link": created.get("link") or "",
            "wp_status": wp_status_actual,
            "rest_base": rest_base,
        }
    except Exception as e:
        return False, str(e), {}


def _maybe_request_gsc_url_inspection(
    project: dict,
    live_url: str,
    wp_status: str | None,
    article_id: str | None,
) -> bool:
    """
    After a live (publish) WordPress post, call Search Console URL Inspection API.
    On a valid inspection response, sets gsc_status to 'inspected' (shown in UI as Inspected).

    Important: Google Search Console's "Request indexing" button is not exposed via a public API
    for general websites. The URL Inspection API returns inspection data; it does not guarantee
    crawling or indexing.
    """
    st = (wp_status or "").strip().lower()
    if st != "publish":
        return False
    url = (live_url or "").strip()
    if not url:
        return False
    proj = dict(project)
    _normalize_project_gsc(proj)
    if not proj.get("gsc_index_on_publish", True):
        return False
    prop = (proj.get("gsc_property_url") or "").strip()
    if not prop:
        return False
    try:
        import google_integration as gi
    except ImportError:
        app.logger.warning("Google integration unavailable (install google-api-python-client et al.).")
        return False
    creds = gi.get_valid_credentials()
    if not creds:
        return False
    try:
        resp = gi.request_url_inspection(creds, prop, url)
        if not gi.gsc_inspection_response_accepted(resp):
            app.logger.warning(
                "Search Console inspect returned no usable inspectionResult for %s (property %s). Not marking GSC inspected.",
                url,
                prop,
            )
            if article_id:
                _update_article_fields(
                    article_id,
                    {
                        "gsc_status": "pending",
                        "gsc_inspection_last_attempt_at": _utc_now_str(),
                        "gsc_inspection_error": "Inspection API returned no inspectionResult.",
                        "gsc_inspection_url": url,
                    },
                )
            return False
        app.logger.info("Search Console URL Inspection accepted for %s (property %s).", url, prop)
        if article_id:
            _update_article_fields(
                article_id,
                {
                    "gsc_status": "inspected",
                    "gsc_inspection_requested_at": _utc_now_str(),
                    "gsc_inspection_last_attempt_at": _utc_now_str(),
                    "gsc_inspection_error": "",
                    "gsc_inspection_url": url,
                },
            )
        return True
    except Exception as e:
        app.logger.warning("Search Console URL Inspection failed for %s: %s", url, e)
        if article_id:
            _update_article_fields(
                article_id,
                {
                    "gsc_status": "pending",
                    "gsc_inspection_last_attempt_at": _utc_now_str(),
                    "gsc_inspection_error": str(e)[:500],
                    "gsc_inspection_url": url,
                },
            )
        return False


@app.post("/projects/<project_id>/articles/<article_id>/wordpress")
def wordpress_post_project(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))

    last = _article_to_last(article)
    if not _is_project_wordpress_configured(project):
        flash("Configure WordPress settings for this project first.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    proj_check = dict(project)
    _normalize_project_image_prompts(proj_check)
    need_featured = len(proj_check.get("image_prompts") or []) > 0
    has_file = os.path.isfile(_article_featured_image_path(article_id))
    allow_without = (request.form.get("confirm_post_without_featured_image") or "").strip() == "1"

    form_body = request.form.get("article_body")
    if not _article_body_matches_stored(form_body, article.get("article")):
        flash("Save your article before posting to WordPress.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    status = (request.form.get("wp_status") or "draft").strip().lower()
    if status not in {"draft", "publish"}:
        status = "draft"

    wp_types, _ = _wp_post_types_for_project(project)
    allowed_bases = {t["rest_base"] for t in wp_types}
    rest_base = _normalize_wp_rest_base(request.form.get("wp_rest_base"), allowed_bases)

    form_site_url = (request.form.get("wp_site_url") or "").strip()
    site_override = form_site_url or None

    ok, err, info = _post_article_to_wordpress(
        project,
        article,
        article_id,
        wp_status=status,
        rest_base_preference=rest_base,
        site_url_override=site_override,
        allow_without_featured_image=allow_without,
    )

    if not ok:
        flash(f"WordPress post failed: {err}", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))

    new_status = _app_article_status_from_wp_rest_status(info.get("wp_status"))
    no_image_suffix = (
        " No featured image was attached (as confirmed)."
        if (not has_file and allow_without)
        else ""
    )
    fk = (request.form.get("wp_focus_keyphrase") or "").strip() or (last.get("focus_keyphrase") or "").strip() or ""
    mt = (request.form.get("wp_meta_title") or "").strip() or (last.get("meta_title") or "").strip() or ""
    md = (request.form.get("wp_meta_description") or "").strip() or (last.get("meta_description") or "").strip() or ""
    posted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    _update_article_fields(
        article_id,
        {
            "status": new_status,
            "wp_post_id": info.get("post_id"),
            "wp_link": info.get("link") or "",
            "wp_rest_base": info.get("rest_base"),
            "focus_keyphrase": fk,
            "meta_title": mt,
            "meta_description": md,
            "posted_at": posted_at,
            "wp_scheduled_at": "",
            "wp_schedule_error": "",
            "wp_schedule_batch_id": "",
            "wp_schedule_batch_index": "",
            "wp_schedule_batch_total": "",
        },
    )

    post_id = info.get("post_id")
    link = info.get("link") or ""
    proj_fresh = _get_project_by_id(project_id) or project
    gsc_requested = _maybe_request_gsc_url_inspection(
        proj_fresh,
        info.get("link") or "",
        info.get("wp_status"),
        article_id,
    )

    if link:
        posted_msg = (
            f"Posted to WordPress ({new_status.capitalize()}). ID {post_id}. Link: {link}{no_image_suffix}"
        )
    else:
        posted_msg = f"Posted to WordPress ({new_status.capitalize()}). ID {post_id}.{no_image_suffix}"
    if gsc_requested:
        posted_msg += (
            " Search Console URL Inspection ran successfully for the live URL. "
            "Indexing status in Google updates on its own schedule."
        )
    flash(posted_msg, "success")
    return redirect(url_for("project_detail", project_id=project_id))


@app.get("/projects/<project_id>/articles/<article_id>/download")
def download_project_article(project_id: str, article_id: str):
    project = _get_project_by_id(project_id)
    article = _get_article_by_id(article_id)
    if not project or not article or (article.get("project_id") or "") != project_id:
        flash("Article not found.", "error")
        return redirect(url_for("home"))
    text = article.get("article") or ""
    if not text.strip():
        flash("Nothing to download yet. Generate an article first.", "error")
        return redirect(url_for("article_edit", project_id=project_id, article_id=article_id))
    filename = _sanitize_filename(article.get("title", "article"))
    return Response(
        text,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/articles")
def articles():
    last = session.get("last_result")
    return render_template("index.html", last=last)


@app.post("/generate")
def generate():
    title = (request.form.get("title") or "").strip()
    keywords_raw = request.form.get("keywords") or ""
    api_key = request.form.get("api_key") or ""
    remember_key = (request.form.get("remember_key") or "") == "on"

    keywords = _parse_keywords(keywords_raw)

    if not title:
        flash("Please enter an article title.", "error")
        return redirect(url_for("articles"))

    if len(keywords) > 10:
        flash("Maximum 10 keywords allowed. Please remove extra keywords.", "error")
        return redirect(url_for("articles"))

    if remember_key and api_key.strip():
        session["api_key"] = api_key.strip()
    else:
        api_key = api_key.strip() or session.get("api_key", "")

    try:
        with ThreadPoolExecutor(max_workers=2) as ex:
            fut_article = ex.submit(_generate_article_markdown, title, keywords, api_key)
            fut_yoast = ex.submit(_generate_yoast_fields, title, keywords, api_key)
            article_md = fut_article.result()
            yoast = fut_yoast.result()
    except Exception as e:
        flash(f"Generation failed: {e}", "error")
        return redirect(url_for("articles"))

    result = {
        "title": title,
        "keywords": keywords,
        "article": (article_md or "").strip(),
        "focus_keyphrase": (yoast.get("focus_keyphrase") or "").strip(),
        "meta_title": (yoast.get("meta_title") or "").strip(),
        "meta_description": (yoast.get("meta_description") or "").strip(),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    session["last_result"] = result
    return render_template("index.html", last=result)


@app.post("/wordpress/post")
def wordpress_post():
    last = session.get("last_result")
    ok, err, info = _execute_wordpress_post_from_last(last or {})
    if not ok:
        flash(f"WordPress post failed: {err}", "error")
        return redirect(url_for("articles"))

    post_id = info.get("post_id")
    link = info.get("link") or ""
    if link:
        flash(f"Posted to WordPress (ID {post_id}). Link: {link}", "success")
    else:
        flash(f"Posted to WordPress (ID {post_id}).", "success")
    return redirect(url_for("articles"))


@app.get("/download")
def download():
    last = session.get("last_result")
    if not last or not last.get("article"):
        flash("Nothing to download yet. Generate an article first.", "error")
        return redirect(url_for("articles"))

    filename = _sanitize_filename(last.get("title", "article"))
    text = last["article"]
    return Response(
        text,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _process_due_scheduled_wordpress_posts() -> None:
    if not _wp_scheduled_processor_lock.acquire(blocking=False):
        return
    try:
        with app.app_context():
            now_utc = datetime.utcnow().replace(microsecond=0)
            now_local = datetime.now().replace(microsecond=0)
            arts = _load_articles()
            candidates: list[dict] = []
            for a in arts:
                sched_utc = _article_wp_scheduled_at_utc_str(a)
                sched = sched_utc or _article_wp_scheduled_at_str(a)
                if not sched:
                    continue
                aid = (a.get("id") or "").strip()
                if not aid:
                    continue
                if a.get("wp_post_id"):
                    continue
                now = now_utc if sched_utc else now_local
                # If the last attempt failed, respect retry delay.
                st = (a.get("wp_schedule_state") or "").strip().lower()
                if st == "error":
                    nr = (a.get("wp_schedule_next_retry_at") or "").strip()
                    if nr:
                        try:
                            dt_nr = datetime.strptime(nr, "%Y-%m-%d %H:%M:%S")
                            if dt_nr > now:
                                continue
                        except ValueError:
                            pass
                try:
                    dt = datetime.strptime(sched, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    _update_article_fields(
                        aid,
                        {
                            "wp_scheduled_at": "",
                            "wp_scheduled_at_utc": "",
                            "wp_schedule_error": "Invalid schedule time stored.",
                        },
                    )
                    continue
                if dt > now:
                    continue
                candidates.append(a)

            candidates.sort(
                key=lambda x: (
                    (x.get("wp_schedule_batch_id") or ""),
                    int(x.get("wp_schedule_batch_index") or 0),
                    _article_wp_scheduled_at_utc_str(x) or _article_wp_scheduled_at_str(x),
                )
            )

            for a in candidates:
                arts = _load_articles()
                aid = (a.get("id") or "").strip()
                cur = _get_article_by_id(aid)
                if not cur or cur.get("wp_post_id"):
                    continue
                sched_utc = _article_wp_scheduled_at_utc_str(cur)
                sched = sched_utc or _article_wp_scheduled_at_str(cur)
                if not sched:
                    continue
                try:
                    dt = datetime.strptime(sched, "%Y-%m-%d %H:%M:%S")
                except ValueError:
                    continue
                now = now_utc if sched_utc else now_local
                if dt > now:
                    continue
                if _earlier_batch_member_blocks(arts, cur, now):
                    continue

                project = _get_project_by_id_unscoped(cur.get("project_id") or "")
                if not project:
                    _update_article_fields(
                        aid,
                        {
                            "wp_scheduled_at": "",
                            "wp_schedule_error": "Project no longer exists.",
                            "wp_schedule_batch_id": "",
                            "wp_schedule_batch_index": "",
                            "wp_schedule_batch_total": "",
                        },
                    )
                    continue
                if not _is_project_wordpress_configured(project):
                    _update_article_fields(
                        aid,
                        {
                            "wp_scheduled_at": "",
                            "wp_schedule_error": "WordPress is not configured for this project.",
                            "wp_schedule_batch_id": "",
                            "wp_schedule_batch_index": "",
                            "wp_schedule_batch_total": "",
                        },
                    )
                    continue
                wp_st = (cur.get("wp_schedule_wp_status") or "draft").strip().lower()
                if wp_st not in ("draft", "publish"):
                    wp_st = "draft"
                proj_check = dict(project)
                _normalize_project_image_prompts(proj_check)
                need_featured = len(proj_check.get("image_prompts") or []) > 0
                fresh = _get_article_by_id(aid) or cur
                # If the article isn't ready yet (still generating), keep it queued.
                if not (fresh.get("article") or "").strip():
                    continue
                if need_featured and not os.path.isfile(_article_featured_image_path(aid)):
                    continue
                _update_article_fields(
                    aid,
                    {
                        "wp_schedule_state": "posting",
                        "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    },
                )
                ok, err, info = _post_article_to_wordpress(
                    project,
                    fresh,
                    aid,
                    wp_status=wp_st,
                    rest_base_preference=None,
                    site_url_override=None,
                    # Scheduled publishing should not silently post without an image when the project uses image prompts.
                    allow_without_featured_image=(not need_featured),
                )
                if not ok:
                    app.logger.warning(
                        "Scheduled WordPress post failed for article %s: %s", aid, err or "unknown"
                    )
                    # Keep the schedule row intact and retry later instead of dropping the item.
                    next_retry = (datetime.now() + timedelta(minutes=5)).strftime("%Y-%m-%d %H:%M:%S")
                    _update_article_fields(
                        aid,
                        {
                            "wp_schedule_state": "error",
                            "wp_schedule_state_updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "wp_schedule_next_retry_at": next_retry,
                            "wp_schedule_error": (err or "WordPress post failed.")[:500],
                        },
                    )
                    continue
                new_status = _app_article_status_from_wp_rest_status(info.get("wp_status"))
                posted_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                _update_article_fields(
                    aid,
                    {
                        "status": new_status,
                        "wp_post_id": info.get("post_id"),
                        "wp_link": info.get("link") or "",
                        "wp_rest_base": info.get("rest_base"),
                        "posted_at": posted_at,
                        "wp_scheduled_at": "",
                        "wp_schedule_error": "",
                        "wp_schedule_batch_id": "",
                        "wp_schedule_batch_index": "",
                        "wp_schedule_batch_total": "",
                        "wp_schedule_state": "",
                        "wp_schedule_state_updated_at": "",
                        "wp_schedule_next_retry_at": "",
                    },
                )
                proj_fresh = _get_project_by_id_unscoped(project.get("id") or "") or project
                _maybe_request_gsc_url_inspection(
                    proj_fresh,
                    info.get("link") or "",
                    info.get("wp_status"),
                    aid,
                )
    except Exception:
        app.logger.exception("Scheduled WordPress posting failed")
    finally:
        _wp_scheduled_processor_lock.release()


def _maybe_trigger_scheduled_wp_posts() -> None:
    """If APScheduler did not run (e.g. Flask reloader on Windows), posting still runs while you browse /projects/."""
    global _wp_bg_trigger_last
    with _wp_bg_trigger_lock:
        t = time.time()
        if t - _wp_bg_trigger_last < 5.0:
            return
        _wp_bg_trigger_last = t
    threading.Thread(target=_process_due_scheduled_wordpress_posts, daemon=True).start()


def _should_run_background_scheduler() -> bool:
    return os.environ.get("WERKZEUG_RUN_MAIN") != "false"


def _start_wp_schedule_scheduler() -> None:
    if not _should_run_background_scheduler():
        app.logger.info(
            "APScheduler skipped (WERKZEUG_RUN_MAIN=false, reloader parent). "
            "Due WordPress posts still run when you open any GET /projects/ URL."
        )
        return
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
    except ImportError:
        app.logger.warning(
            "APScheduler not installed; install APScheduler or rely on browsing /projects/ to post due items."
        )
        return

    sched = BackgroundScheduler(daemon=True)
    sched.add_job(
        _process_due_scheduled_wordpress_posts,
        "interval",
        seconds=60,
        id="wp_scheduled_posts",
        max_instances=1,
        coalesce=True,
    )
    sched.start()
    app.logger.info("APScheduler started: due WordPress posts checked every 60s.")
    import atexit

    atexit.register(lambda: sched.shutdown(wait=False))


@app.before_request
def _trigger_wp_schedule_on_project_get():
    if request.method != "GET":
        return
    p = request.path or ""
    if not p.startswith("/projects/"):
        return
    _maybe_trigger_scheduled_wp_posts()


_start_wp_schedule_scheduler()


if __name__ == "__main__":
    try:
        port = int((os.environ.get("PORT") or "").strip() or "5000")
    except Exception:
        port = 5000
    app.run(host="127.0.0.1", port=port)

