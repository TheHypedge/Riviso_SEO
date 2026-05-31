"""Generate RIVISO_HARDENING_TRACKER.xlsx from the production hardening plan.

Run: backend/.venv/bin/python backend/docs/_build_hardening_tracker.py
Output: backend/docs/RIVISO_HARDENING_TRACKER.xlsx
"""
from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "RIVISO_HARDENING_TRACKER.xlsx")

# ---- palette ----
NAVY = "1F2A44"
BLUE = "2F5597"
LIGHT = "D9E1F2"
ZEBRA = "F2F5FB"
CRIT = "C00000"
HIGH = "ED7D31"
MED = "FFC000"
LOW = "70AD47"
GREY = "808080"

HEADER_FILL = PatternFill("solid", fgColor=BLUE)
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)

WHITE_BOLD = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

SEV_FILL = {
    "Critical": PatternFill("solid", fgColor=CRIT),
    "High": PatternFill("solid", fgColor=HIGH),
    "Medium": PatternFill("solid", fgColor=MED),
    "Low": PatternFill("solid", fgColor=LOW),
    "Info": PatternFill("solid", fgColor=GREY),
}

# Columns: Phase, ID, Category, Item, Severity/Priority, Effort, Evidence (file:line), Fix / Action, Verification, Status, Owner, Notes
COLS = [
    ("Phase", 10),
    ("ID", 9),
    ("Category", 16),
    ("Item", 40),
    ("Severity / Priority", 13),
    ("Effort", 9),
    ("Evidence (file:line)", 34),
    ("Fix / Action", 46),
    ("Verification", 30),
    ("Status", 14),
    ("Owner", 12),
    ("Notes", 22),
]

STATUSES = ["Not Started", "In Progress", "Blocked", "In Review", "Done", "Deferred"]

# ---- rows: (Phase, ID, Category, Item, Sev, Effort, Evidence, Fix, Verify) ----
ROWS = [
    # ===== P0 Critical security =====
    ("P0", "S0.1", "Security", "Fail-fast on placeholder/short SECRET_KEY in production", "Critical", "S",
     "core/config.py:47-49; core/production.py:67-71", "Raise at startup if ENVIRONMENT=production and key is default or <32 chars", "Boot with bad key exits non-zero"),
    ("P0", "S0.2", "Security", "Remove hardcoded admin seed 'Admin@2026'", "Critical", "S",
     "app.py:52-78", "Delete seed; env-gated one-time bootstrap CLI", "grep finds no literal password; fresh DB has no default admin"),
    ("P0", "S0.3", "Security", "Stop returning WordPress app password in API responses", "High", "S",
     "routes/wordpress.py:537,584-585; schemas/project_settings.py:34-36", "Return wp_app_password_set: bool only", "Response JSON contains no secret"),
    ("P0", "S0.4", "Security", "Disallow MONGODB_TLS_INSECURE & OAUTHLIB_INSECURE_TRANSPORT in prod", "High", "S",
     "database.py:88-93; app.py:33-36", "Startup check rejects insecure transport in production", "Prod boot fails if flags set"),
    ("P0", "S0.5", "Security", "Confirm legacy Flask app.py is not deployed/reachable", "High", "M",
     "app.py (whole)", "Remove from deploy target or gate behind disabled flag", "No prod route served by Flask"),
    # ===== P1 Security hardening =====
    ("P1", "S1.1", "Security", "Refresh-token rotation + server-side invalidation", "High", "M",
     "routes/auth.py:700-790", "Issue new RT each refresh; track jti allow/deny list", "Reused old RT is rejected"),
    ("P1", "S1.2", "Security", "Rate-limit /auth/refresh", "High", "S",
     "routes/auth.py:700", "@limiter.limit('20/minute')", "429 after threshold"),
    ("P1", "S1.3", "Security", "Move JWT tokens out of localStorage to httpOnly cookies", "High", "M",
     "frontend/src/lib/api.ts:1038-1060,1282-1347", "Cookie-only auth; remove localStorage", "No tokens in localStorage; XSS cannot read"),
    ("P1", "S1.4", "Security", "Cookie Secure default true in prod + max_age aligned to TTL", "Medium", "S",
     "core/config.py:53; routes/auth.py:127-163,772-788", "Secure+SameSite+expiry", "Cookies have Secure+expiry in prod"),
    ("P1", "S1.5", "Security", "Account lockout after N failed logins", "Medium", "M",
     "routes/auth.py:187", "Per-email backoff/lockout", "Lockout triggers in test"),
    ("P1", "S1.6a", "Security", "SSRF guard on WordPress URL fetch (block private/metadata IPs)", "High", "M",
     "routes/wordpress.py:140-146; services/wordpress_client.py:45-81", "URL allowlist + IP guard; restrict redirects", "Request to 169.254.169.254 / 10.x blocked"),
    ("P1", "S1.6b", "Security", "SSRF guard on Shopify shop resolution", "High", "M",
     "services/shopify_oauth.py:132-137", "Require *.myshopify.com / Admin API only", "Non-shopify host rejected"),
    ("P1", "S1.6c", "Security", "SSRF guard on featured-image + OpenAI ref-image download", "High", "M",
     "services/wordpress_client.py:231-237; services/shopify_article_image.py:47-61; services/openai_client.py:175-178", "HTTPS host allowlist / data-URL only", "Internal IP fetch blocked"),
    ("P1", "S1.7", "Security", "CSRF protection for cookie auth", "Medium", "M",
     "core/deps.py:38-46", "Require X-Requested-With on mutations or SameSite=Strict", "Cross-site POST without header rejected"),
    ("P1", "S1.8", "Security", "Drop localhost origins from production CORS", "Medium", "S",
     "main.py:72-87,234-240", "Env-only strict allowlist in production", "Prod CORS excludes localhost"),
    ("P1", "S1.9", "Security", "Rate-limit expensive endpoints (generate/bulk-upload/publish/research)", "Medium", "M",
     "core/ratelimit.py:13", "Per-user limits on OpenAI-backed routes", "429 after threshold per user"),
    ("P1", "S1.10", "Security", "Trust-proxy config so rate-limit key is not XFF-spoofable", "Medium", "S",
     "core/ratelimit.py:11-13", "Configure trusted proxy or key by user id", "Spoofed XFF does not bypass limit"),
    ("P1", "S1.11", "Security", "Split public liveness from detailed readiness /health", "Medium", "S",
     "routes/health.py:49-60", "Public {status:ok}; detail behind auth", "Anon /health leaks no internals"),
    ("P1", "S1.12", "Security", "Close plan bypasses: humanize + export-consume", "Medium", "M",
     "routes/articles.py:1723; routes/articles.py:901-936", "Add require_plan_action; server-side export gating", "Over-quota humanize/export blocked"),
    ("P1", "S1.13", "Security", "Authenticate WordPress plugin ZIP download", "Medium", "S",
     "routes/wordpress.py:508-527", "Require auth or signed token", "Anon download returns 401"),
    ("P1", "S1.14", "Security", "Add Next.js security headers", "Low", "S",
     "frontend/next.config.ts:15-27", "headers() with HSTS/CSP/X-Frame", "Headers present on frontend responses"),
    # ===== P2 Performance quick wins =====
    ("P2", "P2.1", "Performance", "Request-scoped cache for user/subscription/plan", "High", "M",
     "core/deps.py:58-60; middleware/plan_limits.py:79-90; services/plan_gatekeeper.py:92-97", "Attach to request.state; gatekeeper reads it", "1 user read per request (not 3-5)"),
    ("P2", "P2.2", "Performance", "TTL cache for load_plans()", "Medium", "S",
     "storage.py:1205; services/plan_gatekeeper.py:49-58", "Module cache ~60s invalidated on upsert_plan", "Plans read once per TTL window"),
    ("P2", "P2.3", "Performance", "Wrap all sync storage in run_sync on hot async paths", "High", "M",
     "core/deps.py:58-60; core/project_lookup.py:27-45; routes/wordpress.py:470", "await run_sync(...) everywhere", "No sync PyMongo on event loop (profiler)"),
    ("P2", "P2.4", "Performance", "Add load_articles_by_ids_for_project; drop 20k scans", "High", "M",
     "routes/articles.py:1028-1037; routes/scheduled_jobs.py:246-256", "$in query batch fetch", "Bulk validate uses single $in query"),
    ("P2", "P2.5", "Performance", "asyncio.gather serial pairs (editor-shell, board, shopify sync)", "Medium", "M",
     "routes/articles.py:1248-1262; routes/scheduled_jobs.py:284-295; services/shopify_sync.py:196-250", "Concurrent independent awaits", "Latency = max not sum"),
    ("P2", "P2.6", "Frontend", "Visibility guard + backoff on all poll loops", "Medium", "M",
     "frontend/src/lib/api.ts:788-844; app/projects/[projectId]/page.tsx:1763-1774,3096-3120", "document.hidden check + backoff", "No polling when tab hidden"),
    ("P2", "P2.7", "Frontend", "Replace listArticlesAll waterfall with aggregate endpoint", "High", "M",
     "frontend/src/lib/api.ts:1963-1975; page.tsx:1283-1322", "Use workspaceOverview()", "Overview loads 1 request not 50"),
    ("P2", "P2.8", "Frontend", "Stop shell refetch on tab switch; dedupe GSC analytics", "Medium", "S",
     "page.tsx:1171-1218,1290-1372,2508-2521", "Drop tab from deps; share analytics in state", "Tab switch fires no shell refetch"),
    # ===== P3 Infrastructure =====
    ("P3", "I3.1", "Infrastructure", "Separate worker & scheduler from API processes", "High", "M",
     "main.py:129-138 (lifespan); Procfile", "API ENABLE_SCHEDULER=0/worker=0 + 1 worker + 1 scheduler proc", "API process runs no scheduler/worker"),
    ("P3", "I3.2", "Infrastructure", "Managed MongoDB Atlas M10 (TLS, IP allowlist, backups)", "High", "M",
     "database.py:88-93", "Atlas M10, TLS on, maxIdleTimeMS=30000", "App connects to M10 over TLS"),
    ("P3", "I3.3", "Infrastructure", "Managed Redis (auth + TLS)", "High", "S",
     "services/generation_queue.py", "Single managed Redis with auth/TLS", "Queue uses managed Redis"),
    ("P3", "I3.4", "Infrastructure", "Automated daily backups + tested restore", "High", "S",
     "Atlas / ops", "Continuous backup; quarterly restore drill", "Restore drill succeeds"),
    ("P3", "I3.5", "Infrastructure", "TLS termination + HSTS + gzip at Nginx/Cloudflare", "High", "S",
     "nginx/", "Force HTTPS redirect + HSTS", "HTTP redirects to HTTPS; HSTS header set"),
    ("P3", "I3.6", "Infrastructure", "Connection pool sizing across processes", "Medium", "S",
     "database.py", "Set maxPoolSize per process; sum < Atlas limit", "No connection-limit errors under load"),
    ("P3", "I3.7", "Infrastructure", "Healthchecks + auto-restart (liveness/readiness)", "Medium", "S",
     "routes/health.py", "Container healthcheck -> restart policy", "Killed proc auto-restarts"),
    ("P3", "I3.8", "Infrastructure", "2-instance API behind LB + rolling deploy", "Medium", "M",
     "deploy/", "2 small API instances; zero-downtime deploy", "Deploy with no downtime"),
    ("P3", "I3.9", "Infrastructure", "Secrets via env/secret manager (not files)", "Medium", "S",
     ".env files", "Move to host secret store", "No secrets on disk in prod"),
    ("P3", "I3.10", "Infrastructure", "Rate-limit store in Redis (not in-memory)", "Medium", "S",
     "core/ratelimit.py", "SlowAPI -> Redis storage backend", "Limits consistent across 2 instances"),
    ("P3", "I3.11", "Infrastructure", "Email reliability: queue or native SMTP lib", "Medium", "M",
     "services/email_dispatch.py", "Replace Node subprocess or queue jobs", "Email send retried on failure"),
    # ===== P4 Structural refactor =====
    ("P4", "P4.1", "Performance", "Add Mongo projections (project/user/scheduler light vs full)", "High", "M",
     "storage.py:2323,2490,1515,3683", "Per-call-site projections", "Hot reads exclude heavy fields"),
    ("P4", "P4.2", "Performance", "Route partial writes through $set; batch bulk_update_articles", "High", "M",
     "storage.py:4104-4113,4167-4177", "bulk_write of $set ops", "No full-doc replace on field update"),
    ("P4", "P4.3", "Performance", "Persist has_body + derived listing status", "High", "M",
     "storage.py:3045-3060; routes/articles.py:731-795", "Write-time flags + $match", "Listing avoids body scan"),
    ("P4", "P4.4", "Performance", "Bulk scheduled-job APIs; move heal to worker", "Medium", "M",
     "routes/scheduled_jobs.py:38-78,717-771", "delete_many; background heal", "GET board does no writes"),
    ("P4", "P4.5", "Performance", "Add missing indexes/TTLs (site_maps, monitors, research_cache)", "Medium", "S",
     "database.py:191-229", "create_index + TTL", "explain() shows index use"),
    ("P4", "P4.6", "Architecture", "Typed repositories + domain models (heavy/light fields)", "High", "L",
     "storage.py (whole)", "ArticleRepository/ProjectRepository/etc.", "Business code uses typed models"),
    ("P4", "P4.7", "Architecture", "RequestContext carrying memoized user/project/plan", "Medium", "M",
     "core/deps.py", "DI object", "Single fetch per request entity"),
    ("P4", "P4.8", "Performance", "Expand Motor async coverage for hot reads", "Medium", "L",
     "services/mongo_listings_async.py", "Async reads off thread pool", "Thread pool not saturated under load"),
    # ===== P5 Observability/testing/CI =====
    ("P5", "I5.1", "Observability", "Error tracking (Sentry) API+worker+frontend with PII scrub", "High", "S",
     "—", "Integrate Sentry SDKs", "Test exception appears in Sentry"),
    ("P5", "I5.2", "Observability", "Metrics + dashboards (latency, queue depth, Mongo op time)", "Medium", "M",
     "—", "Prometheus/Grafana or hosted", "Dashboards live"),
    ("P5", "I5.3", "Observability", "Structured logging with request IDs", "Medium", "S",
     "core/logging.py", "Correlation IDs across API/worker", "Logs correlate by request id"),
    ("P5", "I5.4", "CI/CD", "CI: pip-audit/Dependabot + secret scanning (gitleaks)", "High", "S",
     ".github/workflows", "Block merges on Critical", "PR fails on planted secret"),
    ("P5", "I5.5", "CI/CD", "CI: backend pytest + frontend unit tests on PR", "High", "S",
     "backend/tests; frontend", "Green gate to deploy", "PR blocked on test failure"),
    ("P5", "I5.6", "Testing", "Integration tests for auth, plan gating, publish flows", "High", "M",
     "backend/tests", "Cover security-sensitive paths", "Critical flows covered"),
    ("P5", "I5.7", "Observability", "Uptime monitoring + alerting on /health readiness", "Medium", "S",
     "routes/health.py", "External monitor + alerts", "Downtime pages on-call"),
    # ===== P6 Launch readiness =====
    ("P6", "L6.1", "Launch", "Load test @ 15 concurrent users / realistic mix", "High", "M",
     "—", "k6/Locust scenario", "p95 within target, no errors"),
    ("P6", "L6.2", "Launch", "Security re-scan (verify P0/P1 closed)", "High", "S",
     "—", "Re-run audit", "No High/Critical open"),
    ("P6", "L6.3", "Launch", "Runbook + on-call + incident process", "Medium", "S",
     "docs", "Document runbook", "Runbook reviewed"),
    ("P6", "L6.4", "Launch", "Disaster recovery drill (restore from backup)", "High", "S",
     "Atlas", "Restore drill", "RTO/RPO verified"),
    ("P6", "L6.5", "Launch", "Data retention & privacy review (account deletion purges data)", "Medium", "M",
     "routes/profile.py", "Verify deletion end-to-end", "Deleted account leaves no data"),
]

PHASE_TITLES = {
    "P0": "PHASE 0 — Critical Security (stop-the-bleed)",
    "P1": "PHASE 1 — Security Hardening",
    "P2": "PHASE 2 — Performance Quick Wins",
    "P3": "PHASE 3 — Infrastructure for ~50 Users",
    "P4": "PHASE 4 — Structural Refactor (data / OOP layer)",
    "P5": "PHASE 5 — Observability, Testing & CI",
    "P6": "PHASE 6 — Launch Readiness",
}


def style_header_row(ws, row_idx):
    for c in range(1, len(COLS) + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER
        cell.border = BORDER


def build_backlog(wb):
    ws = wb.active
    ws.title = "Backlog"
    ws.sheet_view.showGridLines = False

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
    t = ws.cell(row=1, column=1, value="Riviso — Production Hardening & Scale Tracker (Target ~50 users)")
    t.fill = TITLE_FILL
    t.font = TITLE_FONT
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    # Column widths + header
    for i, (name, width) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    header_row = 2
    for i, (name, _w) in enumerate(COLS, start=1):
        ws.cell(row=header_row, column=i, value=name)
    style_header_row(ws, header_row)
    ws.freeze_panes = "A3"

    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES), allow_blank=True)
    ws.add_data_validation(dv)

    r = header_row + 1
    current_phase = None
    zebra = False
    for (phase, _id, cat, item, sev, eff, evid, fix, verify) in ROWS:
        if phase != current_phase:
            # phase divider row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
            pc = ws.cell(row=r, column=1, value=PHASE_TITLES.get(phase, phase))
            pc.fill = LIGHT_FILL
            pc.font = Font(bold=True, color=NAVY, size=11)
            pc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[r].height = 20
            current_phase = phase
            zebra = False
            r += 1

        values = [phase, _id, cat, item, sev, eff, evid, fix, verify, "Not Started", "", ""]
        for c, val in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if c in (1, 2, 5, 6, 10) else WRAP
            if zebra:
                cell.fill = ZEBRA_FILL
        # severity color
        sev_cell = ws.cell(row=r, column=5)
        if sev in SEV_FILL:
            sev_cell.fill = SEV_FILL[sev]
            sev_cell.font = Font(color="FFFFFF", bold=True)
        # status dropdown
        dv.add(ws.cell(row=r, column=10))
        zebra = not zebra
        r += 1

    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLS))}{r-1}"
    return r


def build_dashboard(wb, last_row):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="Progress Dashboard")
    t.fill = TITLE_FILL
    t.font = TITLE_FONT
    ws.row_dimensions[1].height = 24

    data_range = f"Backlog!$B$3:$B${last_row}"   # ID column (non-empty = real task)
    status_range = f"Backlog!$J$3:$J${last_row}"
    phase_range = f"Backlog!$A$3:$A${last_row}"

    # Status summary
    ws.cell(row=3, column=1, value="By Status").font = BOLD
    ws.cell(row=3, column=2, value="Count").font = BOLD
    rr = 4
    for s in STATUSES:
        ws.cell(row=rr, column=1, value=s).border = BORDER
        ws.cell(row=rr, column=2,
                value=f'=COUNTIF({status_range},"{s}")').border = BORDER
        rr += 1
    ws.cell(row=rr, column=1, value="TOTAL TASKS").font = BOLD
    ws.cell(row=rr, column=2, value=f'=COUNTA({data_range})-COUNTBLANK({data_range})').font = BOLD
    total_row = rr

    # % complete
    ws.cell(row=rr + 2, column=1, value="% Complete").font = BOLD
    ws.cell(row=rr + 2, column=2,
            value=f'=IFERROR(COUNTIF({status_range},"Done")/B{total_row},0)')
    ws.cell(row=rr + 2, column=2).number_format = "0%"

    # By phase
    pstart = rr + 5
    ws.cell(row=pstart, column=1, value="By Phase").font = BOLD
    ws.cell(row=pstart, column=2, value="Total").font = BOLD
    ws.cell(row=pstart, column=3, value="Done").font = BOLD
    for i, p in enumerate(["P0", "P1", "P2", "P3", "P4", "P5", "P6"], start=1):
        row = pstart + i
        ws.cell(row=row, column=1, value=PHASE_TITLES[p]).border = BORDER
        ws.cell(row=row, column=2,
                value=f'=COUNTIFS({phase_range},"{p}",{status_range},"<>")').border = BORDER
        ws.cell(row=row, column=3,
                value=f'=COUNTIFS({phase_range},"{p}",{status_range},"Done")').border = BORDER


def build_legend(wb):
    ws = wb.create_sheet("Legend")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 70
    ws.merge_cells("A1:B1")
    t = ws.cell(row=1, column=1, value="Legend & Conventions")
    t.fill = TITLE_FILL
    t.font = TITLE_FONT

    rows = [
        ("Severity", ""),
        ("Critical", "Exploitable now / secrets exposed — fix before anything else"),
        ("High", "Serious risk or large perf/scale impact — Phase 0/1/2"),
        ("Medium", "Important, bounded impact — schedule within phase"),
        ("Low", "Cleanup / future-proofing"),
        ("", ""),
        ("Effort", ""),
        ("S", "Small — < 1 day"),
        ("M", "Medium — 1-3 days"),
        ("L", "Large — 1 week+"),
        ("", ""),
        ("Status values", ", ".join(STATUSES)),
        ("ID prefix S", "Security item (see RIVISO_PRODUCTION_HARDENING_PLAN.md §A)"),
        ("ID prefix P2/P4", "Performance item (see RIVISO_PERFORMANCE_OPTIMIZATION_AUDIT.md)"),
        ("ID prefix I", "Infrastructure / Ops item"),
        ("ID prefix L", "Launch readiness item"),
        ("Evidence", "Backend paths are relative to backend/app unless app.py/storage.py/database.py (repo root)"),
    ]
    r = 3
    for a, b in rows:
        ca = ws.cell(row=r, column=1, value=a)
        cb = ws.cell(row=r, column=2, value=b)
        cb.alignment = WRAP
        if a in SEV_FILL:
            ca.fill = SEV_FILL[a]
            ca.font = Font(color="FFFFFF", bold=True)
        elif a in ("Severity", "Effort", "Status values"):
            ca.font = BOLD
        r += 1


def main():
    wb = Workbook()
    last_row = build_backlog(wb)
    build_dashboard(wb, last_row)
    build_legend(wb)
    wb.save(OUT)
    print("Wrote", OUT, "with", len(ROWS), "tasks")


if __name__ == "__main__":
    main()
